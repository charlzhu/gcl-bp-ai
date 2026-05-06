from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping

from logistics_excel_mapping import LogisticsExcelFieldMapper, LogisticsExcelRow


@dataclass(frozen=True)
class LogisticsExcelSourceConfig:
    """物流 Excel 来源配置。

    参数：
        year_files：业务年份到 Excel/fixture 文件路径的映射。
    返回值：无。

    业务逻辑：
        P2.2 只允许配置 2023、2024、2025 三个历史年份；2026+ MySQL 不在本轮范围。
    """

    year_files: Mapping[int, Path]

    def get_source_file(self, year: int) -> Path | None:
        """获取指定年份的来源文件。

        参数：
            year：业务年份。
        返回值：
            文件路径；未配置时返回 None。
        """

        path = self.year_files.get(year)
        return Path(path) if path else None


@dataclass(frozen=True)
class LogisticsExcelDataset:
    """物流 Excel 标准化数据集。

    参数：
        year：业务年份。
        source_file：来源文件路径。
        rows：标准化后的物流行列表。
    返回值：无。
    """

    year: int
    source_file: str
    rows: list[LogisticsExcelRow]


class LogisticsExcelWorkbookLoader:
    """物流 Excel loader 基础框架。

    参数：
        config：年份到来源文件路径的配置。
        field_mapper：字段映射器，默认使用 LogisticsExcelFieldMapper。
    返回值：无。

    业务逻辑：
        loader 只负责读取文件和字段标准化，不承担指标计算，避免读取层和口径层耦合。
    """

    def __init__(
        self,
        config: LogisticsExcelSourceConfig,
        field_mapper: LogisticsExcelFieldMapper | None = None,
    ) -> None:
        self.config = config
        self.field_mapper = field_mapper or LogisticsExcelFieldMapper()

    def get_source_file(self, year: int) -> Path | None:
        """获取配置中的来源文件。

        参数：
            year：业务年份。
        返回值：
            文件路径；未配置时返回 None。
        """

        return self.config.get_source_file(year)

    def load_year(self, year: int) -> LogisticsExcelDataset:
        """读取指定年份的 Excel/fixture 数据。

        参数：
            year：业务年份。
        返回值：
            LogisticsExcelDataset。

        业务逻辑：
            基础版支持 xlsx/xlsm/xls 以及测试 fixture 常用的 csv/json。
            真实验收仍通过配置 2023-2025 Excel 文件路径接入。
        """

        source_file = self.config.get_source_file(year)
        if source_file is None:
            raise FileNotFoundError(f"未配置 {year} 年物流 Excel 来源文件。")
        if not source_file.exists():
            raise FileNotFoundError(f"物流 Excel 来源文件不存在: {source_file}")

        raw_rows = self._read_rows(source_file)
        mapped_rows = [
            self.field_mapper.map_row(raw, source_year=year, source_file=source_file, row_no=row_no)
            for row_no, raw in raw_rows
        ]
        return LogisticsExcelDataset(year=year, source_file=str(source_file), rows=mapped_rows)

    def _read_rows(self, source_file: Path) -> list[tuple[int, dict[str, Any]]]:
        """按文件类型读取原始行。

        参数：
            source_file：来源文件路径。
        返回值：
            (行号, 原始行字典) 列表。
        """

        suffix = source_file.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            return self._read_openpyxl_rows(source_file)
        if suffix == ".xls":
            return self._read_pandas_excel_rows(source_file)
        if suffix == ".csv":
            return self._read_csv_rows(source_file)
        if suffix == ".json":
            return self._read_json_rows(source_file)
        raise ValueError(f"暂不支持的物流 Excel fixture 文件类型: {suffix}")

    def _read_openpyxl_rows(self, source_file: Path) -> list[tuple[int, dict[str, Any]]]:
        """读取 xlsx/xlsm 文件。

        参数：
            source_file：来源文件路径。
        返回值：
            (行号, 原始行字典) 列表。
        """

        from openpyxl import load_workbook

        rows: list[tuple[int, dict[str, Any]]] = []
        workbook = load_workbook(source_file, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                iterator = worksheet.iter_rows(values_only=True)
                headers = next(iterator, None)
                if not headers:
                    continue
                header_names = [str(header) if header is not None else "" for header in headers]
                for index, values in enumerate(iterator, start=2):
                    if not values or all(value is None for value in values):
                        continue
                    raw = {header: value for header, value in zip(header_names, values) if header}
                    rows.append((index, raw))
        finally:
            workbook.close()
        return rows

    def _read_pandas_excel_rows(self, source_file: Path) -> list[tuple[int, dict[str, Any]]]:
        """读取旧版 xls 文件。

        参数：
            source_file：来源文件路径。
        返回值：
            (行号, 原始行字典) 列表。
        """

        import pandas as pd

        rows: list[tuple[int, dict[str, Any]]] = []
        excel_file = pd.ExcelFile(source_file)
        for sheet_name in excel_file.sheet_names:
            data_frame = excel_file.parse(sheet_name=sheet_name)
            data_frame = data_frame.where(pd.notnull(data_frame), None)
            for index, row in data_frame.iterrows():
                raw = {str(key): value for key, value in row.to_dict().items()}
                rows.append((int(index) + 2, raw))
        return rows

    def _read_csv_rows(self, source_file: Path) -> list[tuple[int, dict[str, Any]]]:
        """读取 csv fixture。

        参数：
            source_file：来源文件路径。
        返回值：
            (行号, 原始行字典) 列表。
        """

        with source_file.open("r", encoding="utf-8-sig", newline="") as file_obj:
            reader = csv.DictReader(file_obj)
            return [(index, dict(row)) for index, row in enumerate(reader, start=2)]

    def _read_json_rows(self, source_file: Path) -> list[tuple[int, dict[str, Any]]]:
        """读取 json fixture。

        参数：
            source_file：来源文件路径。
        返回值：
            (行号, 原始行字典) 列表。
        """

        payload = json.loads(source_file.read_text(encoding="utf-8"))
        if not isinstance(payload, list):
            raise ValueError("物流 Excel json fixture 必须是对象数组。")
        return [(index, dict(row)) for index, row in enumerate(payload, start=2)]
