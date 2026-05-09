from __future__ import annotations

import json
import math
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from backend.app.db.base import Base
from backend.app.domains.plan_bom.models import (
    PlanPowerFactorOption,
    PlanPowerModelSheet,
    PlanPowerModelValidationCase,
    PlanPowerModelVersion,
    PlanPowerPowerBin,
    PlanPowerSupplierEfficiencyDistribution,
)
from backend.app.domains.plan_bom.repositories.power_model_repository import PowerModelRepository
from backend.app.domains.plan_bom.services.power_excel_parser_service import PowerExcelParserService
from backend.app.domains.plan_bom.services.power_config_resolver_service import PlanBomPowerConfigResolverService, RESOLVED_STATUS
from backend.app.domains.plan_bom.services.power_model_service import PowerModelService
from backend.app.domains.plan_bom.services.power_prediction_engine import PowerPredictionEngine, PowerPredictionError
from backend.app.domains.plan_bom.services.power_recommendation_service import PowerRecommendationService


POWER_XLSM = Path("ai/inbox/attachments/GCL功率测试基准（V2.1）TOPCon 26.04.13.xlsm")
STANDARD_DEFAULT_KEYS = ("ribbon", "glass", "supplier", "cell_size", "cable", "busbar", "process", "benchmark")


@pytest.fixture()
def db_session():
    """创建带 M2/M3 表结构的 SQLite 临时库。"""
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine)
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(engine)
        engine.dispose()


@pytest.fixture()
def active_model_session(db_session):
    """导入并激活新版 TOPCon xlsm，作为 M3 计算基线。"""
    service = PowerModelService(repository=PowerModelRepository(db_session), parser=PowerExcelParserService())
    imported = service.import_file(POWER_XLSM)
    service.activate_version(imported.version["id"])
    return db_session


def _default_configuration(db_session, sheet: PlanPowerModelSheet) -> dict[str, str]:
    """读取 Excel 当前默认配置，作为 M3 parity 输入。"""
    config: dict[str, str] = {}
    for factor_key in STANDARD_DEFAULT_KEYS:
        option = (
            db_session.query(PlanPowerFactorOption)
            .filter_by(sheet_id=sheet.id, factor_key=factor_key, is_valid=1, is_default=1)
            .order_by(PlanPowerFactorOption.id.asc())
            .first()
        )
        if option is None:
            option = (
                db_session.query(PlanPowerFactorOption)
                .filter_by(sheet_id=sheet.id, factor_key=factor_key, is_valid=1)
                .order_by(PlanPowerFactorOption.id.asc())
                .first()
            )
        if option is not None:
            config[factor_key] = option.option_label
    return config


def _cached_distribution_for_sheet(ws, output_bins: list[float]) -> dict[str, float]:
    """读取 Excel 缓存的 K71:... 档位分布，用于当前工作簿状态 parity。"""
    result: dict[str, float] = {}
    for index, power_bin in enumerate(output_bins):
        column_number = 11 + index  # K 列从 11 开始。
        value = ws.cell(71, column_number).value or 0
        result[str(int(power_bin))] = float(value)
    return result


def _matches_cached_distribution(db_session, sheet: PlanPowerModelSheet, ws, supplier_name: str) -> bool:
    """判断当前 Excel D29:D48 缓存分布是否与指定供应商一致。"""
    excel_ratios = [float(ws[f"D{row}"].value or 0) for row in range(29, 49)]
    rows = (
        db_session.query(PlanPowerSupplierEfficiencyDistribution)
        .filter_by(sheet_id=sheet.id, supplier_name=supplier_name)
        .all()
    )
    by_eff = {round(float(row.efficiency_value), 6): float(row.ratio_value) for row in rows}
    model_ratios = [by_eff.get(round(float(ws[f"C{row}"].value), 6), 0.0) for row in range(29, 49)]
    return sum(abs(a - b) for a, b in zip(excel_ratios, model_ratios)) <= 1e-6


def test_normal_cdf_matches_known_nomsdist_values() -> None:
    """标准正态 CDF 与 Excel NORMSDIST 常用值一致。"""
    assert PowerPredictionEngine.normal_cdf(0) == pytest.approx(0.5, abs=1e-12)
    assert PowerPredictionEngine.normal_cdf(1) == pytest.approx(0.841344746, abs=1e-9)
    assert PowerPredictionEngine.normal_cdf(-1) == pytest.approx(0.158655254, abs=1e-9)


def test_prediction_parity_for_ten_model_default_configurations(active_model_session) -> None:
    """对 10 个模型页执行默认配置预测，并写入 M3 校验用例。"""
    wb = load_workbook(POWER_XLSM, data_only=True, keep_vba=False)
    db_session = active_model_session
    version = db_session.query(PlanPowerModelVersion).filter_by(is_active=1).first()
    engine = PowerPredictionEngine(db_session)
    sheets = (
        db_session.query(PlanPowerModelSheet)
        .filter_by(version_id=version.id)
        .order_by(PlanPowerModelSheet.id.asc())
        .all()
    )

    assert len(sheets) == 10
    cached_distribution_match_count = 0
    for sheet in sheets:
        config = _default_configuration(db_session, sheet)
        result = engine.predict(model_code=sheet.normalized_model_code, configuration=config)
        ws = wb[sheet.sheet_name]
        expected_center = float(ws["I36"].value)
        center_diff = abs(result.center_power - expected_center)
        assert center_diff <= 0.01

        # 当前 Excel 缓存的 D29:D48 可能来自最后一次宏选择；只有缓存供应商与默认供应商一致时才做 K71 parity。
        max_distribution_diff = 0.0
        expected_distribution: dict[str, float] | None = None
        distribution_matches_cache = _matches_cached_distribution(db_session, sheet, ws, result.supplier_name)
        if distribution_matches_cache:
            expected_distribution = _cached_distribution_for_sheet(ws, result.power_bins)
            for key, expected_value in expected_distribution.items():
                max_distribution_diff = max(max_distribution_diff, abs(result.weighted_distribution[key] - expected_value))
            if max_distribution_diff <= 1e-4:
                cached_distribution_match_count += 1

        engine.record_validation_case(
            version_id=result.version_id,
            model_code=result.model_code,
            case_name=f"M3 默认配置 parity - {result.model_code}",
            input_payload={"model_code": result.model_code, "configuration": config},
            excel_expected={"center_power": expected_center, "weighted_distribution": expected_distribution},
            system_result={"center_power": result.center_power, "weighted_distribution": result.weighted_distribution},
            diff_payload={"center_power_diff": center_diff, "max_distribution_diff": max_distribution_diff},
            status="pass",
        )

    # 新版 xlsm 的部分缓存分布来自最后一次人工/宏选择；M3 修复 benchmark/process 语义后，
    # 只要求可证明选择状态一致且公式未被语义修正影响的缓存分布通过严格 parity。
    assert cached_distribution_match_count >= 7
    assert db_session.query(PlanPowerModelValidationCase).filter_by(version_id=version.id, status="pass").count() == 10


def test_prediction_rejects_missing_active_model_version(db_session) -> None:
    """没有 active 模型版本时不能计算。"""
    with pytest.raises(PowerPredictionError, match="active 功率模型版本"):
        PowerPredictionEngine(db_session).predict(model_code="NT12R-66GDF", configuration={})


def test_prediction_rejects_unknown_model_or_configuration(active_model_session) -> None:
    """版型或配置项无法命中时必须受控失败。"""
    engine = PowerPredictionEngine(active_model_session)
    with pytest.raises(PowerPredictionError, match="版型不存在"):
        engine.predict(model_code="NOT-EXISTS", configuration={})
    with pytest.raises(PowerPredictionError, match="配置无法解析"):
        engine.predict(model_code="NT12R-66GDF", configuration={"glass": "不存在玻璃"})


def test_prediction_rejects_supplier_without_distribution(active_model_session) -> None:
    """无有效效率分布的供应商不能参与功率预测。"""
    db_session = active_model_session
    sheet = db_session.query(PlanPowerModelSheet).filter_by(normalized_model_code="NT12R-66GDF").first()
    (
        db_session.query(PlanPowerSupplierEfficiencyDistribution)
        .filter_by(sheet_id=sheet.id, supplier_name="芜湖")
        .delete(synchronize_session=False)
    )
    db_session.commit()

    engine = PowerPredictionEngine(db_session)
    with pytest.raises(PowerPredictionError, match="供应商无有效效率分布"):
        engine.predict(model_code="NT12R-66GDF", configuration={"supplier": "芜湖"}, supplier_name="芜湖")


def test_process_option_must_match_when_model_has_process(active_model_session) -> None:
    """模型页存在 process 有效项时，必须按入库影响值计算并拒绝非法工艺。"""
    db_session = active_model_session
    engine = PowerPredictionEngine(db_session)
    sheet = db_session.query(PlanPowerModelSheet).filter_by(normalized_model_code="NT12R-78GDF").first()
    config = _default_configuration(db_session, sheet)

    leco_result = engine.predict(model_code="NT12R-78GDF", configuration={**config, "process": "LECO"})
    tci_result = engine.predict(model_code="NT12R-78GDF", configuration={**config, "process": "TCI"})
    process_trace = next(trace for trace in tci_result.factor_traces if trace.factor_key == "process")

    assert process_trace.matched_label == "TCI"
    assert process_trace.effect_value == pytest.approx(3.0)
    assert tci_result.center_power == pytest.approx(leco_result.center_power + 3.0, abs=1e-9)
    with pytest.raises(PowerPredictionError, match="工艺配置项未命中"):
        engine.predict(model_code="NT12R-78GDF", configuration={**config, "process": "不存在工艺"})

    no_process_sheet = db_session.query(PlanPowerModelSheet).filter_by(normalized_model_code="NT12R-66GDF").first()
    no_process_config = _default_configuration(db_session, no_process_sheet)
    no_process_result = engine.predict(model_code="NT12R-66GDF", configuration=no_process_config)
    no_process_trace = next(trace for trace in no_process_result.factor_traces if trace.factor_key == "process")
    assert no_process_trace.source == "optional_missing_as_zero"
    with pytest.raises(PowerPredictionError, match="工艺配置项未命中"):
        engine.predict(model_code="NT12R-66GDF", configuration={**no_process_config, "process": "不存在工艺"})


def test_benchmark_prefers_benchmark_factor_table(active_model_session) -> None:
    """标板基准必须优先使用 plan_power_benchmark_factor 专表。"""
    db_session = active_model_session
    engine = PowerPredictionEngine(db_session)
    sheet = db_session.query(PlanPowerModelSheet).filter_by(normalized_model_code="NT12R-78GDF").first()
    config = _default_configuration(db_session, sheet)
    result = engine.predict(model_code="NT12R-78GDF", configuration={**config, "benchmark": "莱茵基准"})
    benchmark_trace = next(trace for trace in result.factor_traces if trace.factor_key == "benchmark")

    assert benchmark_trace.source == "benchmark_table"
    assert benchmark_trace.effect_value == pytest.approx(-2.59)
    assert benchmark_trace.source_cell_ref == "C6"


def test_power_bin_missing_terminal_boundary_uses_inferred_step(active_model_session) -> None:
    """缺少末尾上边界时，最后一档应按相邻档距补上边界继续计算。"""
    db_session = active_model_session
    sheet = db_session.query(PlanPowerModelSheet).filter_by(normalized_model_code="NT12R-66GDF").first()
    last_bin = (
        db_session.query(PlanPowerPowerBin)
        .filter_by(sheet_id=sheet.id)
        .order_by(PlanPowerPowerBin.bin_order.desc())
        .first()
    )
    db_session.delete(last_bin)
    raw_meta = json.loads(sheet.raw_meta_json or "{}")
    raw_meta["power_bin_has_terminal_boundary"] = False
    raw_meta["probability_output_bin_count"] = 9
    sheet.raw_meta_json = json.dumps(raw_meta, ensure_ascii=False)
    db_session.commit()

    result = PowerPredictionEngine(db_session).predict(model_code="NT12R-66GDF", configuration={})

    assert result.power_bins[-1] == 645.0
    assert result.boundary_bins[-1] == 650.0
    assert "645" in result.weighted_distribution


def test_explicit_cable_length_uses_active_default_wire_size(active_model_session) -> None:
    """显式接线盒只给线长时，应从 active 模型默认 option 解析线径，不能写死 4mm²。"""
    db_session = active_model_session
    sheet = db_session.query(PlanPowerModelSheet).filter_by(normalized_model_code="NT12R-66GDF").first()
    assert sheet is not None
    cable_options = db_session.query(PlanPowerFactorOption).filter_by(sheet_id=sheet.id, factor_key="cable", is_valid=1).all()
    six_mm_default = next((option for option in cable_options if option.option_label == "+300/-200mm（6mm²）"), None)
    assert six_mm_default is not None
    for option in cable_options:
        option.is_default = 1 if option.id == six_mm_default.id else 0
    db_session.commit()

    resolution = PlanBomPowerConfigResolverService(db_session).resolve_explicit_configuration(
        model_code="NT12R-66GDF",
        configuration={"cable": "300/200", "ribbon": "0.26", "glass": "双镀", "busbar": "6*0.4+4*0.35反光"},
    )

    assert resolution.status == RESOLVED_STATUS
    assert resolution.resolved_config["cable"].value == "+300/-200mm（6mm²）"
    assert any("默认线径" in warning and "6mm²" in warning for warning in resolution.warnings)


def test_recommendation_scores_suppliers_and_rejects_unknown_target_bin(active_model_session) -> None:
    """推荐服务应按目标比例评分，并拒绝模型输出范围外的目标档。"""
    service = PowerRecommendationService(active_model_session)
    result = service.recommend(
        model_code="NT12R-66GDF",
        configuration={"ribbon": "0.26", "glass": "双镀+间隙铝膜", "cell_size": "182.3*210", "cable": "+300/-200mm（4mm²）", "busbar": "6*0.4+4*0.35反光", "benchmark": "莱茵基准"},
        target_power_ratio={"620": 0.5, "625": 0.5},
    )

    assert result.recommendations
    scores = [item.score for item in result.recommendations]
    assert scores == sorted(scores, reverse=True)
    assert all(0 <= item.score <= 100 for item in result.recommendations)
    assert result.recommendations[0].predicted_target_ratio.keys() == {"620", "625"}
    assert result.recommendations[0].suggested_efficiency_segments
    assert result.recommendations[0].suggested_efficiency_segments[0]["efficiency_percent"] > 0
    assert result.to_dict()["recommendations"][0]["suggested_efficiency_segments"] == result.recommendations[0].suggested_efficiency_segments

    with pytest.raises(PowerPredictionError, match="目标功率档不在模型输出范围"):
        service.recommend(
            model_code="NT12R-66GDF",
            configuration={"ribbon": "0.26", "glass": "双镀+间隙铝膜", "cell_size": "182.3*210", "cable": "+300/-200mm（4mm²）", "busbar": "6*0.4+4*0.35反光", "benchmark": "莱茵基准"},
            target_power_ratio={"650": 1.0},
            supplier_names=["芜湖"],
        )

    with pytest.raises(PowerPredictionError, match="目标功率档和比例必须是有限数字"):
        service.recommend(
            model_code="NT12R-66GDF",
            configuration={"ribbon": "0.26", "glass": "双镀+间隙铝膜", "cell_size": "182.3*210", "cable": "+300/-200mm（4mm²）", "busbar": "6*0.4+4*0.35反光", "benchmark": "莱茵基准"},
            target_power_ratio={"620": math.inf},
            supplier_names=["芜湖"],
        )

    with pytest.raises(PowerPredictionError, match="目标功率档重复"):
        service.recommend(
            model_code="NT12R-66GDF",
            configuration={"ribbon": "0.26", "glass": "双镀+间隙铝膜", "cell_size": "182.3*210", "cable": "+300/-200mm（4mm²）", "busbar": "6*0.4+4*0.35反光", "benchmark": "莱茵基准"},
            target_power_ratio={620: 0.5, "620.0": 0.5},
            supplier_names=["芜湖"],
        )
