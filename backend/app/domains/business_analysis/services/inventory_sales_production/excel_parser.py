"""产销存 Excel 只读解析器。

职责边界：
1. 只读取业务提供的 xlsx 文件，不执行宏、不连接外部链接；
2. 将不同年度的宽表统一转换为 DWD 月度事实内存对象；
3. 只导入已发布月份，年度/季度/预算达成率后续由确定性后端服务重算；
4. 不接入问答、不生成 SQL、不让 LLM 直接计算业务数字。
"""

from __future__ import annotations

import calendar
import hashlib
import json
import re
import uuid
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

PARSER_VERSION = "ba_isp_excel_parser_m2_v1"
DOMAIN = "business_analysis"
SUB_DOMAIN = "inventory_sales_production"
MAX_XLSX_BYTES = 50 * 1024 * 1024
MAX_ZIP_MEMBER_COUNT = 300
MAX_ZIP_UNCOMPRESSED_BYTES = 120 * 1024 * 1024

BASE_NAMES = ("合肥", "阜宁", "广德")
MODEL_TYPES = ("182N", "183N", "210N", "210R", "N型", "P型")


@dataclass(slots=True, frozen=True)
class ParsedIspSheet:
    """产销存 Sheet 结构解析结果。

    参数：
        sheet_name: 原始 sheet 名；
        sheet_role: sheet 角色，summary/detail/unknown；
        dimension_ref: Excel 数据范围；
        max_row: 最大行号；
        max_col: 最大列号；
        formula_count: 公式单元格数量；
        merged_cell_count: 合并单元格数量；
        hidden_rows: 隐藏行号列表；
        hidden_cols: 隐藏列字母列表；
        header_rows: 表头行快照。
    """

    sheet_name: str
    sheet_role: str
    dimension_ref: str
    max_row: int
    max_col: int
    formula_count: int
    merged_cell_count: int
    hidden_rows: list[int] = field(default_factory=list)
    hidden_cols: list[str] = field(default_factory=list)
    header_rows: list[list[str | None]] = field(default_factory=list)


@dataclass(slots=True, frozen=True)
class ParsedIspMonthlyFact:
    """产销存月度事实内存对象。

    说明：
        字段与 `dwd_ba_isp_monthly_fact` 基本保持一致，便于仓储层无损落库。
    """

    business_year: int
    business_month: int
    period_label: str
    period_start_date: date
    period_end_date: date
    data_cutoff_month: int
    is_published_month: bool
    metric_code: str
    metric_name: str
    metric_category: str
    aggregation_type: str
    value_decimal: Decimal
    unit_standard: str
    base_name: str | None
    factory_name: str | None
    model_type: str | None
    production_mode: str | None
    trade_scope: str | None
    is_outsourced: bool
    is_consigned: bool
    is_default_external_sales: bool
    source_file_name: str
    source_file_sha256: str
    source_sheet: str
    source_row_index: int
    source_col_index: int
    source_cell_ref: str
    raw_category: str | None
    raw_item: str
    raw_unit: str | None
    parser_version: str
    quality_flags: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True, frozen=True)
class ParsedIspWorkbook:
    """产销存工作簿内存解析结果。"""

    file_name: str
    file_hash: str
    file_size: int
    business_year: int
    data_cutoff_month: int
    source_version_label: str | None
    upload_batch_no: str
    sheet_count: int
    has_vba: bool
    external_link_count: int
    parser_version: str
    quality_status: str
    quality_message: str | None
    quality_flags: list[str]
    sheets: list[ParsedIspSheet] = field(default_factory=list)
    monthly_facts: list[ParsedIspMonthlyFact] = field(default_factory=list)

    @property
    def monthly_fact_count(self) -> int:
        """返回月度事实数量。"""
        return len(self.monthly_facts)


@dataclass(slots=True, frozen=True)
class _MetricMapping:
    """原始行到标准指标和维度的映射结果。"""

    metric_code: str
    metric_name: str
    metric_category: str
    aggregation_type: str
    unit_standard: str = "MW"
    base_name: str | None = None
    factory_name: str | None = None
    model_type: str | None = None
    production_mode: str | None = None
    trade_scope: str | None = None
    is_outsourced: bool = False
    is_consigned: bool = False
    is_default_external_sales: bool = False


class InventorySalesProductionExcelParser:
    """产销存 Excel 到标准月度事实的解析器。

    关键策略：
        1. 通过文件名识别业务年份和截止月份；
        2. 通过表头识别 1-12 月列，只导入 `data_cutoff_month` 以内的月份；
        3. 通过行分类/项目语义归一指标、基地、版型、生产模式、交易范围；
        4. 对 2023 错误年度列、2026 未发布月份等特殊口径写入质量标记。
    """

    def parse_file(self, file_path: str | Path, *, upload_batch_no: str | None = None) -> ParsedIspWorkbook:
        """解析本地 Excel 文件。

        参数：
            file_path: 本地 xlsx 路径；
            upload_batch_no: 可选导入批次号，为空时自动生成。

        返回：
            ParsedIspWorkbook。该对象只在内存中保存解析结果，由仓储负责落库。
        """
        path = Path(file_path)
        content = path.read_bytes()
        return self.parse_bytes(content, file_name=path.name, upload_batch_no=upload_batch_no)

    def parse_bytes(self, content: bytes, *, file_name: str, upload_batch_no: str | None = None) -> ParsedIspWorkbook:
        """解析 Excel 二进制内容。

        参数：
            content: xlsx 文件字节；
            file_name: 原始文件名；
            upload_batch_no: 可选导入批次号。

        返回：
            ParsedIspWorkbook，包含 sheet 结构和 DWD 月度事实。
        """
        self._validate_package(content, file_name=file_name)
        file_hash = hashlib.sha256(content).hexdigest()
        business_year, data_cutoff_month, source_version_label = self._resolve_file_version(file_name)
        quality_flags = self._build_workbook_quality_flags(business_year, data_cutoff_month)
        package_meta = self._inspect_package(content)

        # 说明：公式工作簿用于统计公式数量，缓存工作簿用于读取业务数值。
        formula_wb = load_workbook(PathOrBytes(content), data_only=False, read_only=False)
        cached_wb = load_workbook(PathOrBytes(content), data_only=True, read_only=False)

        parsed_sheets: list[ParsedIspSheet] = []
        monthly_facts: list[ParsedIspMonthlyFact] = []
        for sheet_name in cached_wb.sheetnames:
            cached_ws = cached_wb[sheet_name]
            formula_ws = formula_wb[sheet_name]
            parsed_sheet = self._parse_sheet_structure(formula_ws)
            parsed_sheets.append(parsed_sheet)
            monthly_facts.extend(
                self._parse_sheet_monthly_facts(
                    cached_ws=cached_ws,
                    file_name=file_name,
                    file_hash=file_hash,
                    business_year=business_year,
                    data_cutoff_month=data_cutoff_month,
                    workbook_quality_flags=quality_flags,
                )
            )

        return ParsedIspWorkbook(
            file_name=file_name,
            file_hash=file_hash,
            file_size=len(content),
            business_year=business_year,
            data_cutoff_month=data_cutoff_month,
            source_version_label=source_version_label,
            upload_batch_no=upload_batch_no or self._generate_batch_no(business_year),
            sheet_count=len(cached_wb.sheetnames),
            has_vba=package_meta["has_vba"],
            external_link_count=package_meta["external_link_count"],
            parser_version=PARSER_VERSION,
            quality_status="success",
            quality_message="; ".join(quality_flags) if quality_flags else None,
            quality_flags=quality_flags,
            sheets=parsed_sheets,
            monthly_facts=monthly_facts,
        )

    def _validate_package(self, content: bytes, *, file_name: str) -> None:
        """校验 xlsx 包大小和 ZIP 结构，避免异常文件拖垮解析器。"""
        if not content:
            raise ValueError(f"产销存 Excel 文件为空：{file_name}")
        if len(content) > MAX_XLSX_BYTES:
            raise ValueError(f"产销存 Excel 文件超过大小限制：{file_name}")
        try:
            with ZipFile(PathOrBytes(content)) as archive:
                infos = archive.infolist()
                if len(infos) > MAX_ZIP_MEMBER_COUNT:
                    raise ValueError(f"产销存 Excel ZIP 成员数超过限制：{file_name}")
                uncompressed_size = sum(info.file_size for info in infos)
                if uncompressed_size > MAX_ZIP_UNCOMPRESSED_BYTES:
                    raise ValueError(f"产销存 Excel 解压后大小超过限制：{file_name}")
        except BadZipFile as exc:
            raise ValueError(f"产销存 Excel 不是有效 xlsx 文件：{file_name}") from exc

    def _inspect_package(self, content: bytes) -> dict[str, Any]:
        """读取 xlsx 包元数据。

        返回：
            包含 VBA 是否存在、外部链接数量的字典。
        """
        with ZipFile(PathOrBytes(content)) as archive:
            names = archive.namelist()
        return {
            "has_vba": "xl/vbaProject.bin" in names,
            "external_link_count": sum(1 for name in names if name.startswith("xl/externalLinks/")),
        }

    def _resolve_file_version(self, file_name: str) -> tuple[int, int, str | None]:
        """从文件名识别业务年份和数据截止月。

        规则：
            1. `2026.04` 这类文件名说明只发布到 4 月；
            2. 其他年份文件默认全年 12 个月已发布；
            3. 文件名没有年份时 fail closed，避免误导入。
        """
        year_match = re.search(r"(20\d{2})", file_name)
        if not year_match:
            raise ValueError(f"无法从文件名识别业务年份：{file_name}")
        business_year = int(year_match.group(1))
        version_match = re.search(r"(20\d{2})\.(0?[1-9]|1[0-2])", file_name)
        if version_match:
            cutoff_month = int(version_match.group(2))
            return business_year, cutoff_month, f"{business_year}.{cutoff_month:02d}"
        return business_year, 12, str(business_year)

    def _build_workbook_quality_flags(self, business_year: int, data_cutoff_month: int) -> list[str]:
        """按业务年份生成解析质量标记。"""
        flags: list[str] = []
        if business_year == 2023:
            flags.append("2023 年度列漏 12 月，年度/达成率必须由后端按 1-12 月重算")
        if data_cutoff_month < 12:
            flags.append(f"{business_year} 文件只发布到 {data_cutoff_month} 月，未发布月份不得入库")
        return flags

    def _parse_sheet_structure(self, formula_ws: Worksheet) -> ParsedIspSheet:
        """解析单个 sheet 的结构元数据。"""
        hidden_rows = [index for index, dimension in formula_ws.row_dimensions.items() if dimension.hidden]
        hidden_cols = [column for column, dimension in formula_ws.column_dimensions.items() if dimension.hidden]
        header_values = [self._clean_text(formula_ws.cell(1, col_index).value) for col_index in range(1, formula_ws.max_column + 1)]
        return ParsedIspSheet(
            sheet_name=formula_ws.title,
            sheet_role=self._detect_sheet_role(formula_ws.title),
            dimension_ref=formula_ws.calculate_dimension(),
            max_row=formula_ws.max_row,
            max_col=formula_ws.max_column,
            formula_count=self._count_formulas(formula_ws),
            merged_cell_count=len(formula_ws.merged_cells.ranges),
            hidden_rows=hidden_rows,
            hidden_cols=hidden_cols,
            header_rows=[header_values],
        )

    def _detect_sheet_role(self, sheet_name: str) -> str:
        """识别 sheet 角色。"""
        if "明细" in sheet_name:
            return "detail"
        if sheet_name in {"0103", "2024", "产销存汇总"}:
            return "summary"
        return "unknown"

    def _count_formulas(self, ws: Worksheet) -> int:
        """统计公式单元格数量。"""
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    count += 1
        return count

    def _parse_sheet_monthly_facts(
        self,
        *,
        cached_ws: Worksheet,
        file_name: str,
        file_hash: str,
        business_year: int,
        data_cutoff_month: int,
        workbook_quality_flags: list[str],
    ) -> list[ParsedIspMonthlyFact]:
        """把单个 sheet 的宽表月度列转换为标准月度事实。"""
        month_columns = self._find_month_columns(cached_ws, data_cutoff_month=data_cutoff_month)
        facts: list[ParsedIspMonthlyFact] = []
        current_category: str | None = None
        for row_index in range(2, cached_ws.max_row + 1):
            raw_category, raw_item, raw_unit = self._extract_row_labels(cached_ws, row_index, current_category)
            if raw_category:
                current_category = raw_category
            if not raw_item:
                continue

            mapping = self._normalize_metric(
                business_year=business_year,
                sheet_name=cached_ws.title,
                raw_category=current_category,
                raw_item=raw_item,
            )
            # 预算达成率属于后端计算指标，M2 只沉淀分子/分母月度事实。
            if mapping is None:
                continue

            for month, col_index in month_columns.items():
                raw_value = cached_ws.cell(row_index, col_index).value
                value_decimal = self._to_decimal(raw_value)
                if value_decimal is None:
                    continue
                facts.append(
                    self._build_fact(
                        business_year=business_year,
                        business_month=month,
                        data_cutoff_month=data_cutoff_month,
                        mapping=mapping,
                        value_decimal=value_decimal,
                        file_name=file_name,
                        file_hash=file_hash,
                        sheet_name=cached_ws.title,
                        row_index=row_index,
                        col_index=col_index,
                        raw_category=current_category,
                        raw_item=raw_item,
                        raw_unit=raw_unit,
                        workbook_quality_flags=workbook_quality_flags,
                    )
                )
        return facts

    def _find_month_columns(self, ws: Worksheet, *, data_cutoff_month: int) -> dict[int, int]:
        """根据第 1 行表头定位已发布月份列。"""
        month_columns: dict[int, int] = {}
        for col_index in range(1, ws.max_column + 1):
            header = self._clean_text(ws.cell(1, col_index).value)
            if not header:
                continue
            match = re.fullmatch(r"(\d{1,2})月", header)
            if not match:
                continue
            month = int(match.group(1))
            if 1 <= month <= data_cutoff_month:
                month_columns[month] = col_index
        if not month_columns:
            raise ValueError(f"sheet {ws.title} 未识别到已发布月份列")
        return dict(sorted(month_columns.items()))

    def _extract_row_labels(self, ws: Worksheet, row_index: int, current_category: str | None) -> tuple[str | None, str | None, str | None]:
        """抽取行分类、项目和单位。

        兼容结构：
            1. 2023/2025/2026：A=分类，B=项目，C=单位或月度前列；
            2. 2024 汇总：B=项目，C=基地，D=单位；
            3. 2024 明细：A=类型，B=事项，C=单位。
        """
        headers = [self._clean_text(ws.cell(1, col_index).value) for col_index in range(1, min(ws.max_column, 6) + 1)]
        if headers[:4] == ["业务模块", "项目", "基地", "单位"]:
            category = self._clean_text(ws.cell(row_index, 2).value) or current_category
            item = self._clean_text(ws.cell(row_index, 3).value)
            unit = self._clean_text(ws.cell(row_index, 4).value)
            return category, item, unit
        if headers[:3] == ["类型", "事项", "单位"]:
            category = self._clean_text(ws.cell(row_index, 1).value) or current_category
            item = self._clean_text(ws.cell(row_index, 2).value)
            unit = self._clean_text(ws.cell(row_index, 3).value)
            return category, item, unit

        category = self._clean_text(ws.cell(row_index, 1).value) or current_category
        item = self._clean_text(ws.cell(row_index, 2).value)
        unit = self._clean_text(ws.cell(row_index, 3).value)
        return category, item, unit

    def _normalize_metric(
        self,
        *,
        business_year: int,
        sheet_name: str,
        raw_category: str | None,
        raw_item: str,
    ) -> _MetricMapping | None:
        """将原始行语义归一为标准指标和维度。"""
        del business_year, sheet_name  # 当前 M2 归一主要依赖行语义，年份/sheet 保留给后续扩展。
        category_text = self._compact(raw_category or "")
        item_text = self._compact(raw_item)
        combined = f"{category_text}|{item_text}"

        if "达成率" in item_text:
            return None
        if "寄存" in combined:
            return self._build_consigned_mapping(item_text)
        if any(keyword in combined for keyword in ("库存", "存货", "SAP数据", "SAP-数据")):
            return self._build_inventory_mapping(item_text)
        if "开票" in combined:
            return self._build_invoice_mapping(item_text)
        if "组件事业部" in item_text and "剔除内部交易" in item_text:
            return _MetricMapping(
                metric_code="shipment_external_excluding_internal",
                metric_name="对外销量（剔除内部交易）",
                metric_category="shipment",
                aggregation_type="flow_sum",
                trade_scope="剔除内部交易",
                is_default_external_sales=True,
            )
        if "发货" in combined or "发出量" in combined or "全球营销中心" in item_text:
            return self._build_shipment_mapping(item_text)
        if "目标" in item_text or "预算" in item_text:
            return _MetricMapping(
                metric_code="production_budget",
                metric_name="产量预算/目标",
                metric_category="budget",
                aggregation_type="flow_sum",
                base_name=self._extract_base(item_text),
            )
        if "不含委外" in item_text:
            return _MetricMapping(
                metric_code="production_actual_excluding_oem",
                metric_name="实际产量（不含委外）",
                metric_category="production",
                aggregation_type="flow_sum",
            )
        if "含委外" in item_text or "实际产出" in item_text:
            return _MetricMapping(
                metric_code="production_actual_including_oem",
                metric_name="实际产量（含委外）",
                metric_category="production",
                aggregation_type="flow_sum",
            )
        if "版型" in category_text or self._extract_model_type(item_text):
            return self._build_model_type_mapping(item_text)
        if "委外" in item_text:
            return _MetricMapping(
                metric_code="production_outsourced",
                metric_name="委外加工产量",
                metric_category="production",
                aggregation_type="flow_sum",
                base_name=self._extract_base(item_text),
                production_mode="委外",
                is_outsourced=True,
            )
        if "自产" in item_text:
            return _MetricMapping(
                metric_code="production_self",
                metric_name="自产产量",
                metric_category="production",
                aggregation_type="flow_sum",
                base_name=self._extract_base(item_text),
                production_mode="自产",
            )
        if "代工" in item_text or "OEM" in raw_item or item_text in {"受托加工", "双经销", "TW"}:
            return _MetricMapping(
                metric_code="production_oem",
                metric_name="代工/受托加工产量",
                metric_category="production",
                aggregation_type="flow_sum",
                base_name=self._extract_base(item_text),
                production_mode=self._extract_production_mode(item_text),
                is_outsourced=True,
            )
        if "产量" in combined or "产出" in combined or item_text in {"合计", "合肥", "阜宁", "广德"}:
            return self._build_production_mapping(category_text, item_text)

        # 说明：宽表中可能存在长期生产模式或历史特殊项目，先保留为生产经营量，避免事实丢失。
        return _MetricMapping(
            metric_code="operating_volume_unclassified",
            metric_name="未分类经营数量",
            metric_category="operating",
            aggregation_type="flow_sum",
            base_name=self._extract_base(item_text),
        )

    def _build_inventory_mapping(self, item_text: str) -> _MetricMapping:
        """构造库存/存货指标映射。"""
        base_name = self._extract_base(item_text)
        if base_name:
            return _MetricMapping(
                metric_code="ending_inventory_by_base",
                metric_name="基地期末库存/存货",
                metric_category="inventory",
                aggregation_type="period_end",
                base_name=base_name,
            )
        return _MetricMapping(
            metric_code="ending_inventory_volume",
            metric_name="期末库存/存货",
            metric_category="inventory",
            aggregation_type="period_end",
        )

    def _build_consigned_mapping(self, item_text: str) -> _MetricMapping:
        """构造寄存库存指标映射。"""
        base_name = self._extract_base(item_text)
        if base_name:
            return _MetricMapping(
                metric_code="consigned_inventory_by_base",
                metric_name="基地寄存库存",
                metric_category="consignment",
                aggregation_type="period_end",
                base_name=base_name,
                is_consigned=True,
            )
        return _MetricMapping(
            metric_code="consigned_inventory_volume",
            metric_name="寄存库存",
            metric_category="consignment",
            aggregation_type="period_end",
            is_consigned=True,
        )

    def _build_invoice_mapping(self, item_text: str) -> _MetricMapping:
        """构造开票销量指标映射。"""
        return _MetricMapping(
            metric_code="invoice_sales_volume",
            metric_name="开票销量",
            metric_category="shipment",
            aggregation_type="flow_sum",
            base_name=self._extract_base(item_text),
            trade_scope="开票",
        )

    def _build_shipment_mapping(self, item_text: str) -> _MetricMapping:
        """构造发货/销量指标映射。"""
        base_name = self._extract_base(item_text)
        if "全球营销中心" in item_text:
            return _MetricMapping(
                metric_code="shipment_volume",
                metric_name="发货量/销量",
                metric_category="shipment",
                aggregation_type="flow_sum",
                trade_scope="全球营销中心",
            )
        if base_name:
            return _MetricMapping(
                metric_code="shipment_by_base",
                metric_name="基地发货量/销量",
                metric_category="shipment",
                aggregation_type="flow_sum",
                base_name=base_name,
            )
        return _MetricMapping(
            metric_code="shipment_volume",
            metric_name="发货量/销量",
            metric_category="shipment",
            aggregation_type="flow_sum",
        )

    def _build_model_type_mapping(self, item_text: str) -> _MetricMapping:
        """构造版型产量指标映射。"""
        return _MetricMapping(
            metric_code="production_by_model_type",
            metric_name="版型产量",
            metric_category="production",
            aggregation_type="flow_sum",
            base_name=self._extract_base(item_text),
            model_type=self._extract_model_type(item_text),
        )

    def _build_production_mapping(self, category_text: str, item_text: str) -> _MetricMapping:
        """构造产量/产出指标映射。"""
        base_name = self._extract_base(item_text)
        factory_name = self._extract_factory(item_text)
        if item_text == "合计" or category_text in {"产量", "产出"} and item_text == "合计":
            return _MetricMapping(
                metric_code="production_actual_including_oem",
                metric_name="实际产量（含委外）",
                metric_category="production",
                aggregation_type="flow_sum",
            )
        if base_name:
            return _MetricMapping(
                metric_code="production_by_base",
                metric_name="基地产量",
                metric_category="production",
                aggregation_type="flow_sum",
                base_name=base_name,
                factory_name=factory_name,
            )
        return _MetricMapping(
            metric_code="production_actual_including_oem",
            metric_name="实际产量（含委外）",
            metric_category="production",
            aggregation_type="flow_sum",
        )

    def _build_fact(
        self,
        *,
        business_year: int,
        business_month: int,
        data_cutoff_month: int,
        mapping: _MetricMapping,
        value_decimal: Decimal,
        file_name: str,
        file_hash: str,
        sheet_name: str,
        row_index: int,
        col_index: int,
        raw_category: str | None,
        raw_item: str,
        raw_unit: str | None,
        workbook_quality_flags: list[str],
    ) -> ParsedIspMonthlyFact:
        """构造单条月度事实。"""
        _, last_day = calendar.monthrange(business_year, business_month)
        source_cell_ref = f"{get_column_letter(col_index)}{row_index}"
        fact_quality_flags = {
            "workbook_flags": workbook_quality_flags,
            "source_month_published": business_month <= data_cutoff_month,
        }
        return ParsedIspMonthlyFact(
            business_year=business_year,
            business_month=business_month,
            period_label=f"{business_year}-{business_month:02d}",
            period_start_date=date(business_year, business_month, 1),
            period_end_date=date(business_year, business_month, last_day),
            data_cutoff_month=data_cutoff_month,
            is_published_month=True,
            metric_code=mapping.metric_code,
            metric_name=mapping.metric_name,
            metric_category=mapping.metric_category,
            aggregation_type=mapping.aggregation_type,
            value_decimal=value_decimal,
            unit_standard=mapping.unit_standard,
            base_name=mapping.base_name,
            factory_name=mapping.factory_name,
            model_type=mapping.model_type,
            production_mode=mapping.production_mode,
            trade_scope=mapping.trade_scope,
            is_outsourced=mapping.is_outsourced,
            is_consigned=mapping.is_consigned,
            is_default_external_sales=mapping.is_default_external_sales,
            source_file_name=file_name,
            source_file_sha256=file_hash,
            source_sheet=sheet_name,
            source_row_index=row_index,
            source_col_index=col_index,
            source_cell_ref=source_cell_ref,
            raw_category=raw_category,
            raw_item=raw_item,
            raw_unit=raw_unit,
            parser_version=PARSER_VERSION,
            quality_flags=fact_quality_flags,
        )

    def _to_decimal(self, value: Any) -> Decimal | None:
        """将 Excel 缓存值转换为 Decimal。"""
        if value is None or value == "/":
            return None
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped or stripped == "/":
                return None
            stripped = stripped.replace(",", "")
        else:
            stripped = str(value)
        try:
            return Decimal(stripped)
        except (InvalidOperation, ValueError):
            return None

    def _clean_text(self, value: Any) -> str | None:
        """清洗 Excel 文本值。"""
        if value is None:
            return None
        text = str(value).replace("\u3000", " ").strip()
        text = re.sub(r"\s+", " ", text)
        return text or None

    def _compact(self, value: str) -> str:
        """去掉空白和换行，便于中文关键词判断。"""
        return re.sub(r"\s+", "", value)

    def _extract_base(self, item_text: str) -> str | None:
        """从项目文本中抽取基地名称。"""
        for base_name in BASE_NAMES:
            if base_name in item_text:
                return base_name
        return None

    def _extract_factory(self, item_text: str) -> str | None:
        """从项目文本中抽取工厂名称。"""
        for base_name in BASE_NAMES:
            match = re.search(rf"({base_name}[一二三四五六七八九十]+厂)", item_text)
            if match:
                return match.group(1)
        return None

    def _extract_model_type(self, item_text: str) -> str | None:
        """从项目文本中抽取版型。"""
        for model_type in MODEL_TYPES:
            if model_type in item_text:
                return model_type
        return None

    def _extract_production_mode(self, item_text: str) -> str:
        """从项目文本中抽取生产模式。"""
        if "委外" in item_text:
            return "委外"
        if "代工" in item_text or "OEM" in item_text:
            return "代工"
        if "受托加工" in item_text:
            return "受托加工"
        if "双经销" in item_text:
            return "双经销"
        if item_text == "TW":
            return "TW"
        return "代工"

    def _generate_batch_no(self, business_year: int) -> str:
        """生成导入批次号。"""
        return f"ba_isp_{business_year}_{uuid.uuid4().hex[:12]}"


def PathOrBytes(content: bytes):  # noqa: N802
    """返回新的 BytesIO，确保每个调用方拥有独立读指针。"""
    from io import BytesIO

    return BytesIO(content)


__all__ = [
    "DOMAIN",
    "SUB_DOMAIN",
    "PARSER_VERSION",
    "ParsedIspMonthlyFact",
    "ParsedIspSheet",
    "ParsedIspWorkbook",
    "InventorySalesProductionExcelParser",
]
