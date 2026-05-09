from __future__ import annotations

import json
import re
import warnings
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


FORMULA_POLICY = "semantic_fixed_mode"
MAX_XLSM_BYTES = 50 * 1024 * 1024
MAX_ZIP_MEMBER_COUNT = 500
MAX_ZIP_UNCOMPRESSED_BYTES = 200 * 1024 * 1024

EXPECTED_MODEL_SHEETS = [
    "NT10-72GDF",
    "NT10-78GDF",
    "NT12R-66GDF",
    "NT12R-66GDF (2.0)",
    "NT12-66GDF",
    "NT12R-78GDF ",
    "NT12R-48GDF",
    "NT12R-48BGDF",
    "NT12R-54GDF",
    "NT12R-54BGDF",
]
EXPECTED_SUPPORT_SHEETS = ["标板基准", "更改履历"]
ANCHOR_CELLS = ["A3", "A6", "A9", "A12", "A17", "A20", "A23", "A26", "A76"]
FACTOR_ROWS = {
    3: "ribbon",
    6: "glass",
    9: "supplier",
    12: "cell_size",
    17: "cable",
    20: "busbar",
    23: "process",
    26: "benchmark",
}
PLACEHOLDER_PATTERNS = [
    re.compile(r"^规格\d+$"),
    re.compile(r"^技术\d+$"),
    re.compile(r"^厂家\d+$"),
    re.compile(r"^汇流条\d+$"),
    re.compile(r"^基准\d+$"),
]


@dataclass(slots=True)
class ParsedChangeHistory:
    """内存更改履历。"""

    sequence_no: int | None
    change_content: str
    reviser: str | None
    change_date: str | None
    source_cell_ref: str


@dataclass(slots=True)
class ParsedPowerIssue:
    """内存解析问题项。"""

    level: str
    issue_code: str
    message: str
    source_sheet_name: str | None = None
    source_cell_ref: str | None = None
    raw: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class ParsedFactorOption:
    """内存配置选项。"""

    factor_key: str
    option_label: str
    normalized_option_label: str
    effect_value: Decimal | None
    area_value: Decimal | None
    std_dev_value: Decimal | None
    source_cell_ref: str
    is_default: bool
    is_valid: bool
    invalid_reason: str | None
    raw: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class ParsedSupplierDistribution:
    """内存供应商效率分布。"""

    supplier_name: str
    normalized_supplier_name: str
    efficiency_value: Decimal
    ratio_value: Decimal
    source_cell_ref: str


@dataclass(slots=True)
class ParsedPowerBin:
    """内存功率档位。"""

    power_bin: Decimal
    bin_order: int
    source_cell_ref: str
    is_valid: bool = True


@dataclass(slots=True)
class ParsedBenchmarkFactor:
    """内存标板基准影响项。"""

    model_code: str
    benchmark_name: str
    normalized_benchmark_name: str
    effect_value: Decimal | None
    source_sheet_name: str
    source_cell_ref: str
    raw: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class ParsedPowerModelSheet:
    """内存模型 Sheet 解析结果。"""

    sheet_name: str
    normalized_model_code: str
    cell_count: int | None
    base_power: Decimal | None
    center_power_cell: str
    area_default: Decimal | None
    std_dev_default: Decimal | None
    source_range: str
    raw_meta: dict[str, Any] = field(default_factory=dict)
    factor_options: list[ParsedFactorOption] = field(default_factory=list)
    supplier_distributions: list[ParsedSupplierDistribution] = field(default_factory=list)
    power_bins: list[ParsedPowerBin] = field(default_factory=list)


@dataclass(slots=True)
class ParsedPowerModelWorkbook:
    """内存 Workbook 解析结果。"""

    file_name: str
    file_hash: str
    business_version_label: str | None
    formula_policy: str
    has_vba_project: bool
    vba_project_sha256: str | None
    sheet_count: int
    model_sheet_count: int
    sheets: list[ParsedPowerModelSheet] = field(default_factory=list)
    benchmark_factors: list[ParsedBenchmarkFactor] = field(default_factory=list)
    change_histories: list[ParsedChangeHistory] = field(default_factory=list)
    issues: list[ParsedPowerIssue] = field(default_factory=list)
    parse_summary: dict[str, Any] = field(default_factory=dict)

    @property
    def warning_count(self) -> int:
        """返回 warning 数量。"""
        return sum(1 for issue in self.issues if issue.level == "warning")

    @property
    def error_count(self) -> int:
        """返回 error 数量。"""
        return sum(1 for issue in self.issues if issue.level == "error")


class PowerExcelParserService:
    """功率预测 xlsm 只读解析服务。

    关键边界：
    1. 使用 openpyxl 读取公式与缓存值，不执行 VBA；
    2. 只输出结构化内存结果，不直接落库；
    3. 对异常结构记录 issue，避免静默吞掉 Excel 模板问题。
    """

    def parse_file(self, file_path: str | Path, *, file_hash: str | None = None) -> ParsedPowerModelWorkbook:
        """解析本地 xlsm 文件。

        参数：
            file_path: 本地 xlsm 路径；
            file_hash: 可选文件 SHA256，调用方已计算时可传入。

        返回：
            ParsedPowerModelWorkbook，供版本服务落库。
        """
        path = Path(file_path)
        content = path.read_bytes()
        return self.parse_bytes(content, file_name=path.name, file_hash=file_hash)

    def parse_bytes(self, content: bytes, *, file_name: str, file_hash: str | None = None) -> ParsedPowerModelWorkbook:
        """解析 xlsm 二进制内容。

        参数：
            content: xlsm 文件字节；
            file_name: 原始文件名；
            file_hash: 可选 SHA256，调用方用于幂等防重复。

        返回：
            ParsedPowerModelWorkbook。解析过程不执行任何宏，只检查 VBA 工程是否存在。
        """
        import hashlib

        resolved_hash = file_hash or hashlib.sha256(content).hexdigest()
        issues: list[ParsedPowerIssue] = []
        package_meta = self._inspect_xlsm_package(content)
        has_vba_project = bool(package_meta["has_vba_project"])
        vba_project_sha256 = package_meta["vba_project_sha256"]
        if not has_vba_project:
            issues.append(
                ParsedPowerIssue(
                    level="error",
                    issue_code="VBA_PROJECT_MISSING",
                    message="xlsm 内未发现 xl/vbaProject.bin，无法确认宏模型来源。",
                    raw={"file_name": file_name},
                )
            )

        formula_wb = self._load_workbook(content, data_only=False)
        cached_wb = self._load_workbook(content, data_only=True)
        issues.extend(self._validate_workbook_sheets(formula_wb.sheetnames))

        sheets: list[ParsedPowerModelSheet] = []
        for sheet_name in EXPECTED_MODEL_SHEETS:
            if sheet_name not in formula_wb.sheetnames:
                continue
            formula_ws = formula_wb[sheet_name]
            cached_ws = cached_wb[sheet_name]
            issues.extend(self._validate_anchor_cells(formula_ws))
            model_sheet = self._parse_model_sheet(formula_ws, cached_ws)
            sheets.append(model_sheet)
            issues.extend(self._detect_supplier_title_issues(formula_ws))
            issues.extend(self._detect_semantic_formula_issues(formula_ws))

        benchmark_factors: list[ParsedBenchmarkFactor] = []
        if "标板基准" in formula_wb.sheetnames:
            benchmark_factors = self._parse_benchmark_factors(formula_wb["标板基准"])

        change_histories: list[ParsedChangeHistory] = []
        if "更改履历" in cached_wb.sheetnames:
            change_histories = self._parse_change_histories(cached_wb["更改履历"])

        parse_summary = self._build_parse_summary(
            sheet_count=len(formula_wb.sheetnames),
            model_sheet_count=len(sheets),
            sheets=sheets,
            benchmark_factors=benchmark_factors,
            change_histories=change_histories,
            issues=issues,
        )
        return ParsedPowerModelWorkbook(
            file_name=file_name,
            file_hash=resolved_hash,
            business_version_label=self._infer_business_version_label(file_name),
            formula_policy=FORMULA_POLICY,
            has_vba_project=has_vba_project,
            vba_project_sha256=vba_project_sha256,
            sheet_count=len(formula_wb.sheetnames),
            model_sheet_count=len(sheets),
            sheets=sheets,
            benchmark_factors=benchmark_factors,
            change_histories=change_histories,
            issues=issues,
            parse_summary=parse_summary,
        )

    def _load_workbook(self, content: bytes, *, data_only: bool):
        """加载 Workbook。

        参数：
            content: xlsm 字节；
            data_only: 是否读取 Excel 缓存结果。

        返回：
            openpyxl Workbook。公式版本按要求使用 keep_vba=True + data_only=False。
        """
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            return load_workbook(BytesIO(content), keep_vba=True, data_only=data_only, read_only=False)

    def _inspect_xlsm_package(self, content: bytes) -> dict[str, Any]:
        """检查 xlsm ZIP 包安全边界和 VBA 工程摘要。

        参数：
            content: xlsm 文件字节。

        返回：
            包含 has_vba_project、vba_project_sha256、zip_member_count、zip_uncompressed_bytes 的字典。

        关键业务逻辑：
            M2 只读解析上传文件，不解压到磁盘；这里先限制 ZIP 成员数和未压缩总量，
            避免超大文件或压缩炸弹在 openpyxl 解析阶段造成内存 / CPU DoS。
        """
        import hashlib

        if len(content) > MAX_XLSM_BYTES:
            raise ValueError(f"xlsm 文件大小超过限制：{MAX_XLSM_BYTES} bytes")
        try:
            with ZipFile(BytesIO(content)) as archive:
                infos = archive.infolist()
                if len(infos) > MAX_ZIP_MEMBER_COUNT:
                    raise ValueError(f"xlsm ZIP 成员数超过限制：{MAX_ZIP_MEMBER_COUNT}")
                uncompressed_size = sum(info.file_size for info in infos)
                if uncompressed_size > MAX_ZIP_UNCOMPRESSED_BYTES:
                    raise ValueError(f"xlsm ZIP 未压缩体积超过限制：{MAX_ZIP_UNCOMPRESSED_BYTES} bytes")
                has_vba_project = "xl/vbaProject.bin" in archive.namelist()
                vba_project_sha256 = None
                if has_vba_project:
                    vba_project_sha256 = hashlib.sha256(archive.read("xl/vbaProject.bin")).hexdigest()
                return {
                    "has_vba_project": has_vba_project,
                    "vba_project_sha256": vba_project_sha256,
                    "zip_member_count": len(infos),
                    "zip_uncompressed_bytes": uncompressed_size,
                }
        except BadZipFile as exc:
            raise ValueError("上传文件不是有效的 xlsm/zip 包。") from exc

    def _validate_workbook_sheets(self, sheet_names: list[str]) -> list[ParsedPowerIssue]:
        """校验 Workbook Sheet 集合。

        参数：
            sheet_names: Workbook 实际 Sheet 名称列表。

        返回：
            Sheet 数量、模型页和辅助页缺失问题列表。
        """
        issues: list[ParsedPowerIssue] = []
        expected = EXPECTED_MODEL_SHEETS + EXPECTED_SUPPORT_SHEETS
        if len(sheet_names) != 12:
            issues.append(
                ParsedPowerIssue(
                    level="error",
                    issue_code="SHEET_COUNT_MISMATCH",
                    message=f"Workbook Sheet 数量应为 12，实际为 {len(sheet_names)}。",
                    raw={"sheet_names": sheet_names},
                )
            )
        for sheet_name in expected:
            if sheet_name not in sheet_names:
                issues.append(
                    ParsedPowerIssue(
                        level="error",
                        issue_code="SHEET_MISSING",
                        message=f"缺少必需 Sheet：{sheet_name}",
                        source_sheet_name=sheet_name,
                    )
                )
        return issues

    def _validate_anchor_cells(self, ws: Worksheet) -> list[ParsedPowerIssue]:
        """校验模型页固定锚点。

        参数：
            ws: 模型页 worksheet。

        返回：
            锚点缺失或为空的问题列表。
        """
        issues: list[ParsedPowerIssue] = []
        for cell_ref in ANCHOR_CELLS:
            if self._is_blank(ws[cell_ref].value):
                issues.append(
                    ParsedPowerIssue(
                        level="error",
                        issue_code="ANCHOR_CELL_MISSING",
                        message=f"{ws.title}!{cell_ref} 锚点为空。",
                        source_sheet_name=ws.title,
                        source_cell_ref=cell_ref,
                    )
                )
        return issues

    def _parse_model_sheet(self, formula_ws: Worksheet, cached_ws: Worksheet) -> ParsedPowerModelSheet:
        """解析单个版型模型页。

        参数：
            formula_ws: 读取公式文本的 worksheet；
            cached_ws: 读取 Excel 缓存值的 worksheet。

        返回：
            ParsedPowerModelSheet，包含配置选项、功率档和供应商效率分布。
        """
        power_bin_meta = self._power_bin_meta(formula_ws)
        model_sheet = ParsedPowerModelSheet(
            sheet_name=formula_ws.title,
            normalized_model_code=self._normalize_model_code(formula_ws.title),
            cell_count=self._infer_cell_count(formula_ws.title),
            base_power=self._to_decimal(cached_ws["J1"].value),
            center_power_cell="I36",
            area_default=self._to_decimal(cached_ws["B14"].value),
            std_dev_default=self._to_decimal(cached_ws["B15"].value),
            source_range=formula_ws.calculate_dimension(),
            raw_meta={
                "max_row": formula_ws.max_row,
                "max_column": formula_ws.max_column,
                "base_power_formula": self._cell_raw_value(formula_ws["J1"].value),
                "center_power_cached": self._cell_raw_value(cached_ws["I36"].value),
                "efficiency_start": self._cell_raw_value(cached_ws["C29"].value),
                "center_efficiency": self._cell_raw_value(cached_ws["C36"].value),
                "efficiency_end": self._cell_raw_value(cached_ws["C48"].value),
                "efficiency_step": 0.001,
                "center_row_number": 36,
                "efficiency_first_row_number": 29,
                **power_bin_meta,
                "has_tail_space": formula_ws.title != formula_ws.title.strip(),
            },
        )
        model_sheet.factor_options = self._parse_factor_options(formula_ws, cached_ws)
        model_sheet.power_bins = self._parse_power_bins(formula_ws)
        model_sheet.supplier_distributions = self._parse_supplier_distributions(formula_ws, cached_ws)
        return model_sheet

    def _parse_factor_options(self, formula_ws: Worksheet, cached_ws: Worksheet) -> list[ParsedFactorOption]:
        """解析配置区 A1:Y27 中的配置选项。

        参数：
            formula_ws: 公式 workbook 中的模型页；
            cached_ws: 缓存值 workbook 中的模型页。

        返回：
            配置选项列表。占位项不伪造成有效业务选项，会以 is_valid=False 保留。
        """
        options: list[ParsedFactorOption] = []
        for row_no, factor_key in FACTOR_ROWS.items():
            current_effect = self._to_decimal(cached_ws.cell(row_no, 2).value)
            for col_no in range(3, 26):
                label = formula_ws.cell(row_no, col_no).value
                if self._is_blank(label):
                    continue
                option_label = self._stringify(label)
                effect_value = self._to_decimal(cached_ws.cell(row_no + 1, col_no).value)
                area_value = self._to_decimal(cached_ws.cell(14, col_no).value) if factor_key == "cell_size" else None
                std_dev_value = self._to_decimal(cached_ws.cell(15, col_no).value) if factor_key == "cell_size" else None
                invalid_reason = self._factor_invalid_reason(
                    factor_key=factor_key,
                    option_label=option_label,
                    effect_value=effect_value,
                )
                source_cell_ref = f"{get_column_letter(col_no)}{row_no}"
                options.append(
                    ParsedFactorOption(
                        factor_key=factor_key,
                        option_label=option_label,
                        normalized_option_label=self._normalize_label(option_label),
                        effect_value=effect_value,
                        area_value=area_value,
                        std_dev_value=std_dev_value,
                        source_cell_ref=source_cell_ref,
                        is_default=effect_value is not None and current_effect is not None and effect_value == current_effect,
                        is_valid=invalid_reason is None,
                        invalid_reason=invalid_reason,
                        raw={
                            "sheet_name": formula_ws.title,
                            "anchor_label": self._cell_raw_value(formula_ws.cell(row_no, 1).value),
                            "effect_cell_ref": f"{get_column_letter(col_no)}{row_no + 1}",
                        },
                    )
                )
        return options

    def _parse_power_bins(self, formula_ws: Worksheet) -> list[ParsedPowerBin]:
        """解析功率档位 K28:T28。

        参数：
            formula_ws: 模型页 worksheet。

        返回：
            有效数字功率档列表，48/54 系列会自动跳过“核验”和空列。
        """
        bins: list[ParsedPowerBin] = []
        for col_no in range(11, 21):
            raw_value = formula_ws.cell(28, col_no).value
            power_bin = self._to_decimal(raw_value)
            if power_bin is None:
                continue
            bins.append(
                ParsedPowerBin(
                    power_bin=power_bin,
                    bin_order=len(bins) + 1,
                    source_cell_ref=f"{get_column_letter(col_no)}28",
                )
            )
        return bins

    def _power_bin_meta(self, formula_ws: Worksheet) -> dict[str, Any]:
        """解析功率档概率输出列元数据。

        参数：
            formula_ws: 模型页 worksheet。

        返回：
            包含数字表头数、概率公式输出列数、是否存在末尾上边界的元数据。
        """
        numeric_columns: list[int] = []
        probability_columns: list[int] = []
        for col_no in range(11, 21):
            if self._to_decimal(formula_ws.cell(28, col_no).value) is not None:
                numeric_columns.append(col_no)
                formula_value = formula_ws.cell(29, col_no).value
                if isinstance(formula_value, str) and formula_value.startswith("="):
                    probability_columns.append(col_no)
        return {
            "power_bin_count": len(numeric_columns),
            "probability_output_bin_count": len(probability_columns),
            "power_bin_has_terminal_boundary": len(numeric_columns) > len(probability_columns),
        }

    def _parse_supplier_distributions(self, formula_ws: Worksheet, cached_ws: Worksheet) -> list[ParsedSupplierDistribution]:
        """解析供应商效率分布区 C77:Y96。

        参数：
            formula_ws: 公式 workbook 中的模型页；
            cached_ws: 缓存值 workbook 中的模型页。

        返回：
            有效供应商标题下的效率段比例记录；无效供应商标题不作为有效供应商入库。
        """
        distributions: list[ParsedSupplierDistribution] = []
        for col_no in range(3, 26):
            title = self._resolve_supplier_title(formula_ws, col_no)
            if self._is_blank(title):
                continue
            if self._supplier_title_invalid_reason(title) is not None:
                continue
            supplier_name = self._stringify(title)
            for row_no in range(77, 97):
                ratio_value = self._to_decimal(cached_ws.cell(row_no, col_no).value)
                efficiency_value = self._to_decimal(cached_ws.cell(row_no, 2).value)
                if ratio_value is None or efficiency_value is None:
                    continue
                distributions.append(
                    ParsedSupplierDistribution(
                        supplier_name=supplier_name,
                        normalized_supplier_name=self._normalize_label(supplier_name),
                        efficiency_value=efficiency_value,
                        ratio_value=ratio_value,
                        source_cell_ref=f"{get_column_letter(col_no)}{row_no}",
                    )
                )
        return distributions

    def _detect_supplier_title_issues(self, formula_ws: Worksheet) -> list[ParsedPowerIssue]:
        """识别供应商标题异常。

        参数：
            formula_ws: 模型页 worksheet。

        返回：
            无效供应商标题 warning 列表，例如 0、#REF!、厂家X。
        """
        issues: list[ParsedPowerIssue] = []
        seen: set[str] = set()
        for col_no in range(3, 26):
            title = self._resolve_supplier_title(formula_ws, col_no)
            if self._is_blank(title):
                continue
            invalid_reason = self._supplier_title_invalid_reason(title)
            if invalid_reason is None:
                continue
            source_cell_ref = f"{get_column_letter(col_no)}76"
            key = f"{formula_ws.title}:{source_cell_ref}:{title}"
            if key in seen:
                continue
            seen.add(key)
            issues.append(
                ParsedPowerIssue(
                    level="warning",
                    issue_code="INVALID_SUPPLIER_TITLE",
                    message=f"{formula_ws.title}!{source_cell_ref} 供应商标题无效：{invalid_reason}",
                    source_sheet_name=formula_ws.title,
                    source_cell_ref=source_cell_ref,
                    raw={
                        "title": self._cell_raw_value(title),
                        "row_76_value": self._cell_raw_value(formula_ws.cell(76, col_no).value),
                        "row_9_value": self._cell_raw_value(formula_ws.cell(9, col_no).value),
                    },
                )
            )
        return issues

    def _detect_semantic_formula_issues(self, formula_ws: Worksheet) -> list[ParsedPowerIssue]:
        """识别 R30/R32 语义公式修正项。

        参数：
            formula_ws: 模型页 worksheet。

        返回：
            semantic_fixed_mode 需要追溯的公式 warning。
        """
        if formula_ws.title != "NT12R-66GDF":
            return []
        issues: list[ParsedPowerIssue] = []
        for cell_ref in ["R30", "R32"]:
            formula = formula_ws[cell_ref].value
            if isinstance(formula, str) and re.search(r"L\d+\s*-\s*\$S\$28", formula):
                issues.append(
                    ParsedPowerIssue(
                        level="warning",
                        issue_code="SEMANTIC_FORMULA_FIX_REQUIRED",
                        message=f"{formula_ws.title}!{cell_ref} 疑似仍引用 L行-$S$28，M3 应按 I行-$S$28 语义修正。",
                        source_sheet_name=formula_ws.title,
                        source_cell_ref=cell_ref,
                        raw={"formula": formula, "formula_policy": FORMULA_POLICY},
                    )
                )
        return issues

    def _parse_change_histories(self, ws: Worksheet) -> list[ParsedChangeHistory]:
        """解析更改履历页 A1:D20。

        参数：
            ws: `更改履历` worksheet。

        返回：
            更改履历列表，用于版本追溯。空内容行会跳过，不参与正式计算。
        """
        histories: list[ParsedChangeHistory] = []
        for row_no in range(2, ws.max_row + 1):
            change_content = ws.cell(row_no, 2).value
            if self._is_blank(change_content):
                continue
            histories.append(
                ParsedChangeHistory(
                    sequence_no=int(ws.cell(row_no, 1).value) if self._to_decimal(ws.cell(row_no, 1).value) is not None else None,
                    change_content=self._stringify(change_content),
                    reviser=self._stringify(ws.cell(row_no, 3).value) if not self._is_blank(ws.cell(row_no, 3).value) else None,
                    change_date=self._stringify(ws.cell(row_no, 4).value) if not self._is_blank(ws.cell(row_no, 4).value) else None,
                    source_cell_ref=f"A{row_no}:D{row_no}",
                )
            )
        return histories

    def _parse_benchmark_factors(self, ws: Worksheet) -> list[ParsedBenchmarkFactor]:
        """解析标板基准页 A1:E10。

        参数：
            ws: `标板基准` worksheet。

        返回：
            标板基准影响值列表，包含新版新增的“功率最优”列。
        """
        factors: list[ParsedBenchmarkFactor] = []
        headers = [ws.cell(1, col_no).value for col_no in range(2, 6)]
        for row_no in range(2, 11):
            model_code = ws.cell(row_no, 1).value
            if self._is_blank(model_code):
                continue
            normalized_model_code = self._normalize_model_code(self._stringify(model_code))
            for offset, header in enumerate(headers, start=2):
                if self._is_blank(header):
                    continue
                raw_value = ws.cell(row_no, offset).value
                source_cell_ref = f"{get_column_letter(offset)}{row_no}"
                factors.append(
                    ParsedBenchmarkFactor(
                        model_code=normalized_model_code,
                        benchmark_name=self._stringify(header),
                        normalized_benchmark_name=self._normalize_benchmark_name(self._stringify(header)),
                        effect_value=self._to_decimal(raw_value),
                        source_sheet_name=ws.title,
                        source_cell_ref=source_cell_ref,
                        raw={"value": self._cell_raw_value(raw_value)},
                    )
                )
        return factors

    def _build_parse_summary(
        self,
        *,
        sheet_count: int,
        model_sheet_count: int,
        sheets: list[ParsedPowerModelSheet],
        benchmark_factors: list[ParsedBenchmarkFactor],
        change_histories: list[ParsedChangeHistory],
        issues: list[ParsedPowerIssue],
    ) -> dict[str, Any]:
        """构造解析摘要。

        参数：
            sheet_count: Workbook sheet 总数；
            model_sheet_count: 模型页数量；
            sheets: 模型页解析结果；
            benchmark_factors: 标板解析结果；
            change_histories: 更改履历解析结果；
            issues: 问题列表。

        返回：
            可 JSON 序列化的解析摘要。
        """
        return {
            "sheet_count": sheet_count,
            "model_sheet_count": model_sheet_count,
            "model_codes": [sheet.normalized_model_code for sheet in sheets],
            "factor_option_count": sum(len(sheet.factor_options) for sheet in sheets),
            "valid_factor_option_count": sum(1 for sheet in sheets for option in sheet.factor_options if option.is_valid),
            "power_bin_count": sum(len(sheet.power_bins) for sheet in sheets),
            "supplier_distribution_count": sum(len(sheet.supplier_distributions) for sheet in sheets),
            "benchmark_factor_count": len(benchmark_factors),
            "change_history_count": len(change_histories),
            "latest_change_date": change_histories[-1].change_date if change_histories else None,
            "warning_count": sum(1 for issue in issues if issue.level == "warning"),
            "error_count": sum(1 for issue in issues if issue.level == "error"),
            "formula_policy": FORMULA_POLICY,
        }

    def _resolve_supplier_title(self, ws: Worksheet, col_no: int) -> Any:
        """解析供应商标题。

        参数：
            ws: 模型页 worksheet；
            col_no: 供应商分布区列号。

        返回：
            标题原始值。若第 76 行是 `=C9` 形式，则返回对应第 9 行供应商名。
        """
        title_ref = ws.cell(76, col_no).value
        if isinstance(title_ref, str):
            if "#REF!" in title_ref:
                return title_ref
            match = re.fullmatch(r"=([A-Z]+)9", title_ref.strip())
            if match:
                return ws[f"{match.group(1)}9"].value
        return title_ref

    def _factor_invalid_reason(self, *, factor_key: str, option_label: str, effect_value: Decimal | None) -> str | None:
        """判断配置选项是否为有效业务选项。

        参数：
            factor_key: 配置项 key；
            option_label: 原始选项名；
            effect_value: 影响值。

        返回：
            无效原因；有效时返回 None。
        """
        if self._is_placeholder_label(option_label):
            return "占位选项"
        if factor_key != "process" and effect_value is None:
            return "缺少影响值"
        if factor_key == "process" and effect_value is None and self._is_blank(option_label):
            return "缺少影响值"
        return None

    def _supplier_title_invalid_reason(self, title: Any) -> str | None:
        """判断供应商标题是否有效。

        参数：
            title: 供应商标题原始值。

        返回：
            无效原因；有效时返回 None。
        """
        if isinstance(title, (int, float, Decimal)) and Decimal(str(title)) == Decimal("0"):
            return "供应商标题为 0"
        text = self._stringify(title)
        if "#REF!" in text:
            return "供应商标题为 #REF! 公式错误"
        if re.fullmatch(r"厂家\d+", text):
            return "供应商标题为厂家占位"
        return None

    def _is_placeholder_label(self, value: str) -> bool:
        """判断选项名是否是模板占位。

        参数：
            value: 选项名。

        返回：
            命中规格X、技术X、厂家X、汇流条X、基准X 等占位模式时返回 True。
        """
        return any(pattern.fullmatch(value) for pattern in PLACEHOLDER_PATTERNS)

    def _normalize_model_code(self, value: str) -> str:
        """归一化版型编码。

        参数：
            value: Sheet 名或标板页版型名。

        返回：
            去除首尾空格后的版型编码；保留 `(2.0)` 独立语义。
        """
        return value.strip()

    def _normalize_benchmark_name(self, value: str) -> str:
        """归一化标板名称。

        参数：
            value: Excel 标板列名。

        返回：
            业务确认后的标准标板名称。
        """
        aliases = {
            "北德基准": "新北德",
            "北德": "新北德",
            "莱茵": "莱茵基准",
            "莱茵基准": "莱茵基准",
            "计量院": "中国计量院",
            "中国计量院": "中国计量院",
            "功率最优": "功率最优",
        }
        return aliases.get(value.strip(), value.strip())

    def _normalize_label(self, value: str) -> str:
        """归一化通用标签。

        参数：
            value: 原始标签。

        返回：
            去掉多余空白后的标签。
        """
        return re.sub(r"\s+", "", value.strip())

    def _infer_cell_count(self, sheet_name: str) -> int | None:
        """从版型编码推断电池片数量。

        参数：
            sheet_name: 模型 Sheet 名。

        返回：
            GDF 前的片数；无法识别时返回 None。
        """
        match = re.search(r"-(\d+)[A-Z]*GDF", sheet_name.strip())
        if not match:
            return None
        return int(match.group(1))

    def _infer_business_version_label(self, file_name: str) -> str | None:
        """从文件名推断业务版本标签。

        参数：
            file_name: 原始文件名。

        返回：
            例如 `TOPCon 26.04.13`；无法识别时返回 None。
        """
        match = re.search(r"(TOPCon)\s*(\d{2}\.\d{2}\.\d{2})", file_name, flags=re.IGNORECASE)
        if match:
            return f"{match.group(1)} {match.group(2)}"
        return None

    def _to_decimal(self, value: Any) -> Decimal | None:
        """将单元格值转 Decimal。

        参数：
            value: openpyxl 单元格值。

        返回：
            数字值返回 Decimal；空、公式、文本和错误值返回 None。
        """
        if self._is_blank(value):
            return None
        if isinstance(value, str) and value.startswith("="):
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None

    def _is_blank(self, value: Any) -> bool:
        """判断单元格值是否为空。

        参数：
            value: 任意单元格值。

        返回：
            None 或空白字符串返回 True。
        """
        return value is None or (isinstance(value, str) and value.strip() == "")

    def _stringify(self, value: Any) -> str:
        """稳定转成字符串。

        参数：
            value: 任意单元格值。

        返回：
            去掉首尾空白后的字符串。
        """
        return str(value).strip()

    def _cell_raw_value(self, value: Any) -> Any:
        """把单元格值转成 JSON 可序列化值。

        参数：
            value: 单元格原始值。

        返回：
            Decimal 等不可直接序列化对象会转成字符串。
        """
        if isinstance(value, Decimal):
            return str(value)
        return value


def dumps_power_json(payload: Any) -> str:
    """将功率模型解析 payload 序列化为 JSON。

    参数：
        payload: 任意可被 json.dumps 处理的对象，Decimal 会转字符串。

    返回：
        UTF-8 中文友好的 JSON 字符串。
    """
    return json.dumps(payload, ensure_ascii=False, default=str)


__all__ = [
    "FORMULA_POLICY",
    "MAX_XLSM_BYTES",
    "EXPECTED_MODEL_SHEETS",
    "ParsedBenchmarkFactor",
    "ParsedChangeHistory",
    "ParsedFactorOption",
    "ParsedPowerBin",
    "ParsedPowerIssue",
    "ParsedPowerModelSheet",
    "ParsedPowerModelWorkbook",
    "ParsedSupplierDistribution",
    "PowerExcelParserService",
    "dumps_power_json",
]
