from __future__ import annotations

import hashlib
import re
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from fastapi import UploadFile
from openpyxl import load_workbook

from backend.app.domains.plan_bom.constants import (
    MATERIAL_CATEGORY_BUSBAR,
    MATERIAL_CATEGORY_GAP_FILM,
    MATERIAL_CATEGORY_GLASS,
    MATERIAL_CATEGORY_INTERCONNECT_BAR,
    MATERIAL_CATEGORY_JUNCTION_BOX,
    MATERIAL_CATEGORY_OTHER,
    PLAN_BOM_NOISE_DESCRIPTION_KEYWORDS,
    PLAN_BOM_NOISE_NAME_KEYWORDS,
    PLAN_BOM_NOISE_SAP_CODE_KEYWORDS,
    PLAN_BOM_SECTION_STOP_KEYWORDS,
    SOURCE_TAG_MANUAL_IMPORT,
    SOURCE_TYPE_EXCEL,
    STATUS_FAILED,
    STATUS_SUCCESS,
)
from backend.app.domains.plan_bom.identity import build_file_instance_key, build_order_identity_key
from backend.app.domains.plan_bom.models import (
    PlanBomHeader,
    PlanBomImportBatch,
    PlanBomMaterialLine,
    PlanBomRevision,
)
from backend.app.domains.plan_bom.repositories.import_repository import PlanBomImportRepository
from backend.app.domains.plan_bom.schemas.import_excel import PlanBomImportIssue, PlanBomImportReport


@dataclass(frozen=True)
class _SheetRows:
    """Excel 单个 sheet 的二维数据。

    参数：
        name: sheet 名称；
        rows: 原始单元格二维数组，行号按 Excel 1 基开始映射。
    """

    name: str
    rows: list[list[Any]]


@dataclass
class _ParsedSheet:
    """单个 sheet 解析结果。"""

    header: PlanBomHeader | None
    material_lines: list[PlanBomMaterialLine]
    revisions: list[PlanBomRevision]


@dataclass
class _RevisionFallback:
    """修订区兜底识别结果。

    说明：
        部分真实 Excel 只有“修订版本 / 修订内容 / 修订人 / 生效日期”表头，
        但下一行直接进入“编制/日期”，没有显式修订明细行。
        这类文件不能直接判定为“无生效日期”，需要把审批日期和备注区内容
        作为开发期兜底线索回填到修订结果中，保证当前版本排序链路可用。
    """

    effective_date: date | None
    reviser: str | None
    revision_content: str | None
    raw_row_no: int | None


class PlanBomExcelImportService:
    """计划 BOM Excel 导入服务。

    职责边界：
    1. 读取 xlsx/xls 文件；
    2. 解析 BOM 头、材料行、修订区；
    3. 检测材料行唯一键冲突；
    4. 写入批次和解析结果；
    5. 不实现查询、导出业务逻辑、前端或 SAP 接入。
    """

    HEADER_ALIASES = {
        "file_no": {"文件号", "文件编号", "bom文件号"},
        "version_no": {"版本号", "版本", "bom版本"},
        "order_no": {"订单号", "订单编码", "单号"},
        "order_name": {"订单名称", "订单名", "项目名称", "型号"},
    }
    MATERIAL_ALIASES = {
        "line_no": {"序号", "行号"},
        "sap_code": {"sap编码", "sap码", "物料编码", "物料号"},
        "material_name": {"物料名称", "材料名称", "名称"},
        "description": {"描述", "规格描述", "物料描述"},
        "standard_usage": {"标准用量", "用量"},
        "unit": {"单位", "计量单位"},
        "production_loss": {"生产损耗", "损耗"},
        "remark": {"备注", "说明"},
        "order_no": {"订单号", "订单编码"},
        "version_no": {"版本号", "版本"},
    }
    REVISION_ALIASES = {
        "revision_version": {"修订版本", "版本", "版本号"},
        "revision_content": {"修订内容", "变更内容", "内容"},
        "reviser": {"修订人", "修改人"},
        "effective_date": {"生效日期", "日期"},
    }
    SECTION_STOP_KEYWORDS = PLAN_BOM_SECTION_STOP_KEYWORDS
    NOISE_NAME_KEYWORDS = PLAN_BOM_NOISE_NAME_KEYWORDS
    NOISE_DESCRIPTION_KEYWORDS = PLAN_BOM_NOISE_DESCRIPTION_KEYWORDS
    NOISE_SAP_CODE_KEYWORDS = PLAN_BOM_NOISE_SAP_CODE_KEYWORDS

    def __init__(self, repository: PlanBomImportRepository) -> None:
        self.repository = repository

    def list_upload_history(self, *, limit: int = 50) -> list[dict[str, Any]]:
        """查询 BOM Excel 上传历史。

        参数：
            limit: 最多返回的历史批次数量。

        返回：
            批次摘要字典列表，按上传时间倒序排列。
        """
        safe_limit = max(1, min(limit, 200))
        return self.repository.list_batches(limit=safe_limit)

    async def import_upload(self, file: UploadFile, *, batch_id: str | None = None) -> PlanBomImportReport:
        """导入 FastAPI 上传的 Excel 文件。

        参数：
            file: FastAPI 上传文件；
            batch_id: 可选批次号，为空时自动生成。

        返回：
            导入结果报告。
        """
        content = await file.read()
        return self.import_bytes(content, file_name=file.filename or "plan_bom.xlsx", batch_id=batch_id)

    def import_file(self, file_path: str | Path, *, batch_id: str | None = None) -> PlanBomImportReport:
        """从本地路径导入 Excel 文件。

        参数：
            file_path: Excel 文件路径；
            batch_id: 可选批次号，便于测试或外部系统指定。

        返回：
            导入结果报告。
        """
        path = Path(file_path)
        content = path.read_bytes()
        return self.import_bytes(content, file_name=path.name, batch_id=batch_id)

    def import_bytes(self, content: bytes, *, file_name: str, batch_id: str | None = None) -> PlanBomImportReport:
        """从字节内容导入 Excel 文件。

        参数：
            content: Excel 文件二进制内容；
            file_name: 文件名，用于识别扩展名和报告展示；
            batch_id: 可选批次号。

        返回：
            导入结果报告。
        """
        actual_batch_id = batch_id or self._generate_batch_id()
        file_hash = hashlib.sha256(content).hexdigest()
        errors: list[PlanBomImportIssue] = []
        warnings: list[PlanBomImportIssue] = []
        sheets = self._load_sheets(content, file_name=file_name)

        headers: list[PlanBomHeader] = []
        material_lines: list[PlanBomMaterialLine] = []
        revisions: list[PlanBomRevision] = []
        # 说明：
        # 真实 BOM 会在同一 sheet 中重复出现同一个 SAP 编码的材料行。
        # 为了区分“完全重复”“仅用量不同”“真实内容冲突”三类情况，
        # 这里保留当前批次已接受的材料行对象，后续按需做就地合并。
        material_line_registry: dict[tuple[str, str, str, str], PlanBomMaterialLine] = {}

        for sheet in sheets:
            parsed_sheet = self._parse_sheet(
                sheet=sheet,
                batch_id=actual_batch_id,
                file_name=file_name,
                file_hash=file_hash,
                errors=errors,
                warnings=warnings,
                material_line_registry=material_line_registry,
            )
            if parsed_sheet.header:
                headers.append(parsed_sheet.header)
            material_lines.extend(parsed_sheet.material_lines)
            revisions.extend(parsed_sheet.revisions)

        if not headers:
            errors.append(
                PlanBomImportIssue(
                    level="error",
                    stage="PARSE_HEADER",
                    message="未解析到任何有效 BOM 头，无法完成批次入库。",
                )
            )

        status = STATUS_FAILED if errors else STATUS_SUCCESS
        batch = PlanBomImportBatch(
            batch_id=actual_batch_id,
            source_type=SOURCE_TYPE_EXCEL,
            source_tag=SOURCE_TAG_MANUAL_IMPORT,
            file_name=file_name,
            file_hash=file_hash,
            status=status,
            total_files=1,
            total_headers=len(headers),
            total_lines=len(material_lines),
            error_message="; ".join(issue.message for issue in errors[:5]) if errors else None,
            finished_at=datetime.now(),
        )
        report = PlanBomImportReport(
            batch_id=actual_batch_id,
            status=status,
            file_name=file_name,
            file_hash=file_hash,
            sheet_count=len(sheets),
            header_count=len(headers),
            material_line_count=len(material_lines),
            revision_count=len(revisions),
            error_count=len(errors),
            warning_count=len(warnings),
            rollback_applied=status == STATUS_FAILED,
            persisted_business_data=False,
            errors=errors,
            warnings=warnings,
        )
        if status == STATUS_FAILED:
            self.repository.save_batch_only(batch=batch)
            return report

        try:
            self.repository.save_import_result(
                batch=batch,
                headers=headers,
                material_lines=material_lines,
                revisions=revisions,
            )
            report.persisted_business_data = True
            return report
        except Exception as exc:
            errors.append(
                PlanBomImportIssue(
                    level="error",
                    stage="PERSIST_IMPORT",
                    message=f"批次入库失败，已整批回滚：{exc}",
                )
            )
            batch.status = STATUS_FAILED
            batch.error_message = "; ".join(issue.message for issue in errors[:5])
            self.repository.save_batch_only(batch=batch)
            report.status = STATUS_FAILED
            report.error_count = len(errors)
            report.rollback_applied = True
            report.persisted_business_data = False
            report.errors = errors
            return report

    def _parse_sheet(
        self,
        *,
        sheet: _SheetRows,
        batch_id: str,
        file_name: str,
        file_hash: str,
        errors: list[PlanBomImportIssue],
        warnings: list[PlanBomImportIssue],
        material_line_registry: dict[tuple[str, str, str, str], PlanBomMaterialLine],
    ) -> _ParsedSheet:
        """解析单个 sheet 的 BOM 头、材料行和修订区。"""
        header_values = self._parse_header_values(sheet.rows)
        material_header_index, material_columns = self._find_table_header(sheet.rows, self.MATERIAL_ALIASES, {"sap_code", "material_name"})
        if material_header_index is not None and (not header_values.get("order_no") or not header_values.get("version_no")):
            self._backfill_header_from_material_rows(sheet.rows, material_header_index, material_columns, header_values)

        order_no = self._clean_text(header_values.get("order_no"))
        version_no = self._clean_text(header_values.get("version_no"))
        order_name = self._clean_text(header_values.get("order_name"))
        if not order_no or not version_no:
            errors.append(
                PlanBomImportIssue(
                    level="error",
                    stage="PARSE_HEADER",
                    sheet_name=sheet.name,
                    message="缺少订单号或版本号，跳过该 sheet。",
                    raw_payload={"header": header_values},
                )
            )
            return _ParsedSheet(header=None, material_lines=[], revisions=[])

        order_identity_key = build_order_identity_key(order_no, order_name, file_name)
        file_instance_key = build_file_instance_key(
            order_identity_key,
            version_no,
            SOURCE_TYPE_EXCEL,
            file_name,
            file_hash,
        )
        revisions = self._parse_revisions(
            sheet=sheet,
            order_no=order_no,
            version_no=version_no,
            order_identity_key=order_identity_key,
            file_instance_key=file_instance_key,
            batch_id=batch_id,
            material_header_index=material_header_index,
        )
        effective_date = self._select_header_effective_date(revisions)
        header = PlanBomHeader(
            order_no=order_no,
            version_no=version_no,
            order_identity_key=order_identity_key,
            file_instance_key=file_instance_key,
            file_no=self._clean_text(header_values.get("file_no")),
            order_name=order_name,
            effective_date=effective_date,
            source_type=SOURCE_TYPE_EXCEL,
            source_tag=SOURCE_TAG_MANUAL_IMPORT,
            import_batch_id=batch_id,
            raw_file_name=file_name,
            raw_sheet_name=sheet.name,
        )
        material_lines = self._parse_material_lines(
            sheet=sheet,
            material_header_index=material_header_index,
            material_columns=material_columns,
            header=header,
            batch_id=batch_id,
            errors=errors,
            warnings=warnings,
            material_line_registry=material_line_registry,
        )
        if material_header_index is None:
            warnings.append(
                PlanBomImportIssue(
                    level="warning",
                    stage="PARSE_MATERIAL",
                    sheet_name=sheet.name,
                    message="未识别到材料明细表头，该 sheet 只入库 BOM 头和修订区。",
                    key=f"{order_no}|{version_no}",
                )
            )
        return _ParsedSheet(header=header, material_lines=material_lines, revisions=revisions)

    def _parse_material_lines(
        self,
        *,
        sheet: _SheetRows,
        material_header_index: int | None,
        material_columns: dict[str, int],
        header: PlanBomHeader,
        batch_id: str,
        errors: list[PlanBomImportIssue],
        warnings: list[PlanBomImportIssue],
        material_line_registry: dict[tuple[str, str, str, str], PlanBomMaterialLine],
    ) -> list[PlanBomMaterialLine]:
        """解析材料行并执行同批次唯一键冲突检测。

        关键逻辑：
        1. 真实 Excel 的修订区、备注区、图纸区通常都在材料区下方；
        2. 一旦识别到这些区块头，必须立即停止材料解析，不能继续向下吞行；
        3. 否则会把备注说明、图纸行、审批信息误入材料表，进而污染查询结果。
        """
        if material_header_index is None:
            return []

        material_lines: list[PlanBomMaterialLine] = []
        for row_index in range(material_header_index + 1, len(sheet.rows)):
            row = sheet.rows[row_index]
            if self._is_blank_row(row):
                continue
            if self._is_material_section_boundary(row):
                break

            order_no = self._clean_text(self._get_by_columns(row, material_columns, "order_no")) or header.order_no
            version_no = self._clean_text(self._get_by_columns(row, material_columns, "version_no")) or header.version_no
            sap_code = self._clean_text(self._get_by_columns(row, material_columns, "sap_code"))
            material_name = self._clean_text(self._get_by_columns(row, material_columns, "material_name"))
            description = self._clean_text(self._get_by_columns(row, material_columns, "description"))
            remark = self._clean_text(self._get_by_columns(row, material_columns, "remark"))
            if not sap_code and not material_name:
                continue
            row_no = row_index + 1
            if self._is_noise_material(sap_code=sap_code, material_name=material_name, description=description):
                continue
            if not sap_code:
                errors.append(
                    PlanBomImportIssue(
                        level="error",
                        stage="MATERIAL_REQUIRED_FIELD",
                        sheet_name=sheet.name,
                        row_no=row_no,
                        message="材料行缺少 SAP 编码，已跳过该行。",
                        key=f"{order_no}|{version_no}|<empty>",
                        raw_payload=self._row_payload(row),
                    )
                )
                continue
            if not material_name:
                errors.append(
                    PlanBomImportIssue(
                        level="error",
                        stage="MATERIAL_REQUIRED_FIELD",
                        sheet_name=sheet.name,
                        row_no=row_no,
                        message="材料行缺少物料名称，已跳过该行。",
                        key=f"{order_no}|{version_no}|{sap_code}",
                        raw_payload=self._row_payload(row),
                    )
                )
                continue

            standard_usage = self._parse_decimal(self._get_by_columns(row, material_columns, "standard_usage"))
            unit = self._clean_text(self._get_by_columns(row, material_columns, "unit"))
            production_loss = self._clean_text(self._get_by_columns(row, material_columns, "production_loss"))
            material_category = self._classify_material(sap_code, material_name, description)
            key = (header.file_instance_key, version_no, sap_code, SOURCE_TYPE_EXCEL)
            issue_key = f"{order_no}|{version_no}|{sap_code}|{SOURCE_TYPE_EXCEL}"
            line = PlanBomMaterialLine(
                order_no=order_no,
                version_no=version_no,
                order_identity_key=header.order_identity_key,
                file_instance_key=header.file_instance_key,
                sap_code=sap_code,
                line_no=self._clean_text(self._get_by_columns(row, material_columns, "line_no")),
                material_name=material_name,
                material_category=material_category,
                description=description,
                standard_usage=standard_usage,
                unit=unit,
                production_loss=production_loss,
                remark=remark,
                replacement_marker=self._detect_replacement_marker(material_name, description, remark),
                source_type=SOURCE_TYPE_EXCEL,
                source_tag=SOURCE_TAG_MANUAL_IMPORT,
                import_batch_id=batch_id,
                raw_row_no=row_no,
            )
            if key in material_line_registry:
                previous_line = material_line_registry[key]
                if self._same_material_signature(previous_line, line):
                    warnings.append(
                        PlanBomImportIssue(
                            level="warning",
                            stage="MATERIAL_DUPLICATE",
                            sheet_name=sheet.name,
                            row_no=row_no,
                            message="材料行唯一键完全重复，已保留第一条并跳过该行。",
                            key=issue_key,
                            raw_payload=self._row_payload(row),
                        )
                    )
                elif self._can_merge_usage_variant(previous_line, line):
                    self._merge_material_usage_variant(previous_line, line)
                    warnings.append(
                        PlanBomImportIssue(
                            level="warning",
                            stage="MATERIAL_USAGE_VARIANT",
                            sheet_name=sheet.name,
                            row_no=row_no,
                            message="同一 SAP 编码仅用量不同，已按较大用量合并并保留一条。",
                            key=issue_key,
                            raw_payload=self._row_payload(row),
                        )
                    )
                elif previous_line.material_category == MATERIAL_CATEGORY_OTHER and material_category == MATERIAL_CATEGORY_OTHER:
                    warnings.append(
                        PlanBomImportIssue(
                            level="warning",
                            stage="MATERIAL_AUX_CONFLICT",
                            sheet_name=sheet.name,
                            row_no=row_no,
                            message="非核心辅材存在同键重复，已保留第一条并跳过该行。",
                            key=issue_key,
                            raw_payload=self._row_payload(row),
                        )
                    )
                else:
                    errors.append(
                        PlanBomImportIssue(
                            level="error",
                            stage="MATERIAL_CONFLICT",
                            sheet_name=sheet.name,
                            row_no=row_no,
                            message="材料行唯一键冲突但字段内容不同，已跳过该行。",
                            key=issue_key,
                            raw_payload=self._row_payload(row),
                        )
                    )
                continue
            material_line_registry[key] = line
            material_lines.append(line)
        return material_lines

    def _parse_revisions(
        self,
        *,
        sheet: _SheetRows,
        order_no: str,
        version_no: str,
        order_identity_key: str,
        file_instance_key: str,
        batch_id: str,
        material_header_index: int | None,
    ) -> list[PlanBomRevision]:
        """解析修订区记录。

        关键逻辑：
        1. 真实 BOM 的修订区通常位于材料区下方，而不是材料区上方；
        2. 旧逻辑在发现修订区位于材料区下方时，会因为 `row_index >= material_header_index`
           直接提前退出，导致修订区整块被跳过；
        3. 这里改为按“修订区起点 -> 下一个审批区/图纸区/材料区边界”来截取。
        """
        revision_header_index, revision_columns = self._find_table_header(
            sheet.rows,
            self.REVISION_ALIASES,
            {"revision_version", "effective_date"},
        )
        if revision_header_index is None:
            return []

        revisions: list[PlanBomRevision] = []
        revision_fallback = self._build_revision_fallback(
            rows=sheet.rows,
            revision_header_index=revision_header_index,
        )
        revision_stop_index = self._find_revision_stop_index(
            rows=sheet.rows,
            revision_header_index=revision_header_index,
            material_header_index=material_header_index,
        )
        for row_index in range(revision_header_index + 1, revision_stop_index):
            row = sheet.rows[row_index]
            if self._is_blank_row(row):
                continue
            revision_version = self._clean_text(
                self._get_span_value(
                    row=row,
                    columns=revision_columns,
                    field_name="revision_version",
                )
            )
            revision_content = self._clean_text(self._get_by_columns(row, revision_columns, "revision_content"))
            effective_date = self._parse_date(self._get_by_columns(row, revision_columns, "effective_date"))
            reviser = self._clean_text(self._get_by_columns(row, revision_columns, "reviser"))
            if not revision_version and not revision_content and not effective_date:
                continue
            revisions.append(
                PlanBomRevision(
                    order_no=order_no,
                    version_no=version_no,
                    order_identity_key=order_identity_key,
                    file_instance_key=file_instance_key,
                    revision_version=revision_version,
                    revision_content=revision_content,
                    reviser=reviser,
                    effective_date=effective_date,
                    source_type=SOURCE_TYPE_EXCEL,
                    source_tag=SOURCE_TAG_MANUAL_IMPORT,
                    import_batch_id=batch_id,
                    raw_row_no=row_index + 1,
                )
            )
        if revisions:
            return revisions

        # 说明：
        # 某些真实文件只有修订表头，没有显式修订明细行，下一行直接进入审批区。
        # 这时用审批日期和上一行备注做开发期兜底，避免 effective_date 长期缺失。
        if revision_fallback and revision_fallback.effective_date:
            revisions.append(
                PlanBomRevision(
                    order_no=order_no,
                    version_no=version_no,
                    order_identity_key=order_identity_key,
                    file_instance_key=file_instance_key,
                    revision_version=version_no,
                    revision_content=revision_fallback.revision_content,
                    reviser=revision_fallback.reviser,
                    effective_date=revision_fallback.effective_date,
                    source_type=SOURCE_TYPE_EXCEL,
                    source_tag=SOURCE_TAG_MANUAL_IMPORT,
                    import_batch_id=batch_id,
                    raw_row_no=revision_fallback.raw_row_no,
                )
            )
        return revisions

    def _load_sheets(self, content: bytes, *, file_name: str) -> list[_SheetRows]:
        """读取 Excel 文件并返回所有 sheet 的二维数组。"""
        suffix = Path(file_name).suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            workbook = load_workbook(BytesIO(content), data_only=True)
            return [
                _SheetRows(
                    name=sheet.title,
                    rows=[[cell for cell in row] for row in sheet.iter_rows(values_only=True)],
                )
                for sheet in workbook.worksheets
            ]
        if suffix == ".xls":
            excel_file = pd.ExcelFile(BytesIO(content))
            sheets: list[_SheetRows] = []
            for sheet_name in excel_file.sheet_names:
                frame = excel_file.parse(sheet_name=sheet_name, header=None)
                frame = frame.where(pd.notnull(frame), None)
                sheets.append(_SheetRows(name=sheet_name, rows=frame.values.tolist()))
            return sheets
        raise ValueError("计划 BOM 导入当前仅支持 xlsx/xlsm/xls 文件")

    def _parse_header_values(self, rows: list[list[Any]]) -> dict[str, Any]:
        """从 sheet 中解析键值对形式的 BOM 头字段。"""
        values: dict[str, Any] = {}
        normalized_aliases = self._normalized_aliases(self.HEADER_ALIASES)
        for row_index, row in enumerate(rows):
            for column_index, cell_value in enumerate(row):
                label = self._normalize_label(cell_value)
                if not label:
                    continue
                for field_name, aliases in normalized_aliases.items():
                    if label not in aliases or values.get(field_name):
                        continue
                    values[field_name] = self._next_value(rows, row_index, column_index)
        return values

    def _backfill_header_from_material_rows(
        self,
        rows: list[list[Any]],
        material_header_index: int,
        material_columns: dict[str, int],
        header_values: dict[str, Any],
    ) -> None:
        """当 BOM 头键值对缺失时，从材料表行级列补齐订单号和版本号。"""
        for row in rows[material_header_index + 1 :]:
            if self._is_blank_row(row):
                continue
            if not header_values.get("order_no"):
                header_values["order_no"] = self._get_by_columns(row, material_columns, "order_no")
            if not header_values.get("version_no"):
                header_values["version_no"] = self._get_by_columns(row, material_columns, "version_no")
            if header_values.get("order_no") and header_values.get("version_no"):
                return

    def _find_table_header(
        self,
        rows: list[list[Any]],
        aliases: dict[str, set[str]],
        required_fields: set[str],
    ) -> tuple[int | None, dict[str, int]]:
        """查找表头行并返回字段到列下标的映射。"""
        normalized_aliases = self._normalized_aliases(aliases)
        for row_index, row in enumerate(rows):
            columns: dict[str, int] = {}
            for column_index, value in enumerate(row):
                label = self._normalize_label(value)
                if not label:
                    continue
                for field_name, field_aliases in normalized_aliases.items():
                    if label in field_aliases and field_name not in columns:
                        columns[field_name] = column_index
            if required_fields.issubset(columns):
                return row_index, columns
        return None, {}

    def _find_revision_stop_index(
        self,
        *,
        rows: list[list[Any]],
        revision_header_index: int,
        material_header_index: int | None,
    ) -> int:
        """确定修订区的结束位置。

        说明：
        - 当修订区位于材料区上方时，到材料表头为止；
        - 当修订区位于材料区下方时，到审批区、图纸区或下一个结构化区块为止。
        """
        if material_header_index is not None and revision_header_index < material_header_index:
            return material_header_index

        for row_index in range(revision_header_index + 1, len(rows)):
            row = rows[row_index]
            if self._looks_like_approval_header(row) or self._looks_like_drawing_header(row):
                return row_index
            material_like_header_index, _ = self._find_table_header(rows[row_index : row_index + 1], self.MATERIAL_ALIASES, {"sap_code", "material_name"})
            if material_like_header_index == 0:
                return row_index
        return len(rows)

    def _build_revision_fallback(
        self,
        *,
        rows: list[list[Any]],
        revision_header_index: int,
    ) -> _RevisionFallback | None:
        """构造无显式修订行时的修订兜底结果。

        说明：
            目标样本里存在“修订表头 -> 审批表头 -> 审批日期行”的结构。
            这类文件虽然没有正式修订明细，但审批日期和备注描述已经足够支撑
            开发期的当前版本判定，因此这里补一条兜底修订记录。
        """
        remark_row = rows[revision_header_index - 1] if revision_header_index - 1 >= 0 else []
        approval_header_row = rows[revision_header_index + 1] if revision_header_index + 1 < len(rows) else []
        approval_value_row = rows[revision_header_index + 2] if revision_header_index + 2 < len(rows) else []

        if not self._looks_like_approval_header(approval_header_row):
            return None

        approval_pairs = self._extract_approval_date_pairs(approval_value_row)
        if not approval_pairs:
            return None

        latest_pair = max(approval_pairs, key=lambda item: item[1])
        revision_content = self._extract_note_content(remark_row)
        return _RevisionFallback(
            effective_date=latest_pair[1],
            reviser=latest_pair[0],
            revision_content=revision_content,
            raw_row_no=revision_header_index + 3,
        )

    @staticmethod
    def _extract_approval_date_pairs(row: list[Any]) -> list[tuple[str | None, date]]:
        """从审批日期行中提取“姓名 / 日期”对。

        说明：
            审批行常见格式为“樊 娜/2026-03-20”，也可能只出现日期。
            这里只取能解析出日期的单元格，并优先保留日期最大的那一项。
        """
        pairs: list[tuple[str | None, date]] = []
        for value in row:
            text = PlanBomExcelImportService._clean_text(value)
            if not text:
                continue
            match = re.search(r"(20\d{2}[-/.]\d{1,2}[-/.]\d{1,2})", text)
            if not match:
                continue
            parsed_date = PlanBomExcelImportService._parse_date(match.group(1))
            if not parsed_date:
                continue
            reviser_text = text[: match.start()].rstrip("/ ").strip() or None
            pairs.append((reviser_text, parsed_date))
        return pairs

    @staticmethod
    def _extract_note_content(row: list[Any]) -> str | None:
        """提取备注行中的业务描述，用作兜底修订内容。"""
        if not row:
            return None
        cleaned_values = [PlanBomExcelImportService._clean_text(value) for value in row]
        non_empty_values = [value for value in cleaned_values if value]
        if len(non_empty_values) >= 2 and non_empty_values[0] == "备注":
            return non_empty_values[1]
        if len(non_empty_values) >= 3 and non_empty_values[1] == "备注":
            return non_empty_values[2]
        return None

    @staticmethod
    def _same_material_signature(previous_line: PlanBomMaterialLine, current_line: PlanBomMaterialLine) -> bool:
        """判断两条材料行是否完全一致。"""
        return (
            previous_line.material_category == current_line.material_category
            and PlanBomExcelImportService._normalize_material_compare_name(previous_line.material_name)
            == PlanBomExcelImportService._normalize_material_compare_name(current_line.material_name)
            and previous_line.description == current_line.description
            and previous_line.standard_usage == current_line.standard_usage
            and previous_line.unit == current_line.unit
            and previous_line.production_loss == current_line.production_loss
            and previous_line.remark == current_line.remark
        )

    @staticmethod
    def _can_merge_usage_variant(previous_line: PlanBomMaterialLine, current_line: PlanBomMaterialLine) -> bool:
        """判断两条材料行是否只在用量层面存在差异。

        说明：
            中样本失败文件里，同一个 SAP 编码会在“搭配虚拟件”和“实际材料”区段重复出现。
            这类行的物料名称、描述、单位和损耗都一致，只有标准用量不同。
            继续把它当成致命冲突会导致整批回滚，因此这里降级为“用量变体合并”。
        """
        return (
            previous_line.material_category == current_line.material_category
            and PlanBomExcelImportService._normalize_material_compare_name(previous_line.material_name)
            == PlanBomExcelImportService._normalize_material_compare_name(current_line.material_name)
            and previous_line.description == current_line.description
            and previous_line.unit == current_line.unit
            and previous_line.production_loss == current_line.production_loss
            and previous_line.remark == current_line.remark
            and previous_line.standard_usage != current_line.standard_usage
        )

    @staticmethod
    def _merge_material_usage_variant(previous_line: PlanBomMaterialLine, current_line: PlanBomMaterialLine) -> None:
        """合并同一 SAP 编码的用量变体，保留较大用量和较早原始行号。"""
        previous_usage = previous_line.standard_usage or Decimal("0")
        current_usage = current_line.standard_usage or Decimal("0")
        if current_usage > previous_usage:
            previous_line.standard_usage = current_line.standard_usage
        if previous_line.raw_row_no and current_line.raw_row_no:
            previous_line.raw_row_no = min(previous_line.raw_row_no, current_line.raw_row_no)
        elif current_line.raw_row_no:
            previous_line.raw_row_no = current_line.raw_row_no

    @staticmethod
    def _normalize_material_compare_name(material_name: str | None) -> str | None:
        """标准化冲突比较用的物料名称。

        说明：
            真实 BOM 会把同一材料写成“互联条”“互联条 1”“互联条 2”。
            这类尾号主要是版式分组标记，不应单独导致同 SAP 编码冲突升级为失败。
        """
        text = PlanBomExcelImportService._clean_text(material_name)
        if not text:
            return text
        return re.sub(r"\s+\d+$", "", text).strip()

    @staticmethod
    def _normalized_aliases(aliases: dict[str, set[str]]) -> dict[str, set[str]]:
        """标准化字段别名。"""
        return {field_name: {PlanBomExcelImportService._normalize_label(alias) for alias in field_aliases} for field_name, field_aliases in aliases.items()}

    @staticmethod
    def _normalize_label(value: Any) -> str:
        """标准化表头或键名，降低空格、冒号、大小写差异影响。"""
        if value is None:
            return ""
        text = str(value).strip().lower()
        return re.sub(r"[\s:：_/-]+", "", text)

    @staticmethod
    def _clean_text(value: Any) -> str | None:
        """清理 Excel 单元格文本。"""
        if value is None:
            return None
        if isinstance(value, float) and pd.isna(value):
            return None
        text = str(value).strip()
        if not text or text.lower() == "nan":
            return None
        return text

    @staticmethod
    def _is_blank_row(row: list[Any]) -> bool:
        """判断是否为空行。"""
        return all(PlanBomExcelImportService._clean_text(value) is None for value in row)

    @staticmethod
    def _get_by_columns(row: list[Any], columns: dict[str, int], field_name: str) -> Any:
        """按字段列映射读取行值。"""
        column_index = columns.get(field_name)
        if column_index is None or column_index >= len(row):
            return None
        return row[column_index]

    @staticmethod
    def _get_span_value(row: list[Any], columns: dict[str, int], field_name: str) -> Any:
        """读取表头所在合并区块中的第一个非空值。

        说明：
        真实 BOM 的修订区表头常出现合并单元格，标题在首列，但实际值落在后续列。
        这里按“当前字段列 -> 下一个字段列之前”的范围兜底查第一个非空值。
        """
        column_index = columns.get(field_name)
        if column_index is None:
            return None

        next_columns = sorted(index for name, index in columns.items() if name != field_name and index > column_index)
        stop_index = next_columns[0] if next_columns else len(row)
        for current_index in range(column_index, min(stop_index, len(row))):
            value = row[current_index]
            if PlanBomExcelImportService._clean_text(value) is not None:
                return value
        return None

    @staticmethod
    def _next_value(rows: list[list[Any]], row_index: int, column_index: int) -> Any:
        """读取键值对右侧或下方的第一个非空值。"""
        row = rows[row_index]
        for next_column in range(column_index + 1, len(row)):
            value = PlanBomExcelImportService._clean_text(row[next_column])
            if value is not None:
                return row[next_column]
        if row_index + 1 < len(rows) and column_index < len(rows[row_index + 1]):
            return rows[row_index + 1][column_index]
        return None

    @staticmethod
    def _row_payload(row: list[Any]) -> dict[str, Any]:
        """生成原始行快照。"""
        return {f"col_{index + 1}": value for index, value in enumerate(row)}

    @staticmethod
    def _parse_decimal(value: Any) -> Decimal | None:
        """解析标准用量数值。"""
        text = PlanBomExcelImportService._clean_text(value)
        if text is None:
            return None
        try:
            return Decimal(text.replace(",", ""))
        except (InvalidOperation, ValueError):
            return None

    @staticmethod
    def _parse_date(value: Any) -> date | None:
        """解析 Excel 日期。"""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = PlanBomExcelImportService._clean_text(value)
        if text is None:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.date()

    @staticmethod
    def _classify_material(sap_code: str, material_name: str, description: str | None) -> str:
        """按物料名称和描述归类 5 类核心材料。

        说明：
        真实 BOM 中存在大量“图纸行、标签行、虚拟件、备注行”混在材料区附近。
        这些行即使文本里包含“玻璃、接线盒、焊带”等关键词，也不应进入核心材料结果。
        """
        if PlanBomExcelImportService._is_noise_material(
            sap_code=sap_code,
            material_name=material_name,
            description=description,
        ):
            return MATERIAL_CATEGORY_OTHER
        text = f"{material_name or ''} {description or ''}"
        if "玻璃" in text:
            return MATERIAL_CATEGORY_GLASS
        if "间隙膜" in text or "间隙贴膜" in text or "贴膜" in text:
            return MATERIAL_CATEGORY_GAP_FILM
        if "互联条" in text or "焊带" in text:
            return MATERIAL_CATEGORY_INTERCONNECT_BAR
        if "汇流条" in text or "汇流" in text:
            return MATERIAL_CATEGORY_BUSBAR
        if "接线盒" in text or "线盒" in text:
            return MATERIAL_CATEGORY_JUNCTION_BOX
        return MATERIAL_CATEGORY_OTHER

    @staticmethod
    def _detect_replacement_marker(*values: str | None) -> str | None:
        """检测替代料显式标识，只做原样提示，不做关系推断。"""
        text = " ".join(value or "" for value in values)
        if "替代" in text or "代用" in text:
            return "explicit_replacement_text"
        return None

    @staticmethod
    def _select_header_effective_date(revisions: list[PlanBomRevision]) -> date | None:
        """从修订区中选出 BOM 头可用的最新生效日期。"""
        dates = [revision.effective_date for revision in revisions if revision.effective_date]
        return max(dates) if dates else None

    @staticmethod
    def _looks_like_revision_header(row: list[Any]) -> bool:
        """判断当前行是否像修订区表头，用于材料解析跳过。"""
        labels = {PlanBomExcelImportService._normalize_label(value) for value in row}
        return "修订版本" in labels and ("生效日期" in labels or "修订内容" in labels)

    @classmethod
    def _is_material_section_boundary(cls, row: list[Any]) -> bool:
        """判断当前行是否已进入材料区后的其他区块。

        说明：
        一旦进入备注区、修订区、审批区或图纸区，材料解析必须立即停止，
        否则真实样本中的长备注和图纸行会被误吞进材料区。
        """
        return (
            cls._looks_like_revision_header(row)
            or cls._looks_like_note_row(row)
            or cls._looks_like_approval_header(row)
            or cls._looks_like_drawing_header(row)
        )

    @classmethod
    def _looks_like_note_row(cls, row: list[Any]) -> bool:
        """判断当前行是否像备注区或注释区。"""
        non_empty_values = [cls._clean_text(value) for value in row if cls._clean_text(value)]
        if not non_empty_values:
            return False
        first_value = non_empty_values[0]
        if any(first_value.startswith(prefix) for prefix in ("注：", "注:", "注")):
            return True
        if first_value == "备注":
            return True
        if len(non_empty_values) >= 2 and non_empty_values[1] == "备注":
            return True
        return False

    @classmethod
    def _looks_like_approval_header(cls, row: list[Any]) -> bool:
        """判断当前行是否像审批区表头。"""
        labels = {cls._clean_text(value) for value in row if cls._clean_text(value)}
        return {"编制/日期", "审核/日期"}.issubset(labels)

    @classmethod
    def _looks_like_drawing_header(cls, row: list[Any]) -> bool:
        """判断当前行是否像图纸区表头。"""
        labels = {cls._clean_text(value) for value in row if cls._clean_text(value)}
        return {"文控文件号", "文控版本", "图纸名称"}.issubset(labels)

    @classmethod
    def _is_noise_material(
        cls,
        *,
        sap_code: str | None,
        material_name: str | None,
        description: str | None,
    ) -> bool:
        """判断当前材料行是否属于核心材料查询噪音。

        说明：
        这里专门处理两类真实误判：
        1. 虚拟件描述里带“玻璃”等关键词；
        2. 图纸、标签、印字说明文本里带“接线盒”等关键词。
        """
        clean_name = material_name or ""
        clean_desc = description or ""
        clean_sap_code = sap_code or ""
        if clean_sap_code == "备注":
            return True
        if any(keyword in clean_name for keyword in cls.NOISE_NAME_KEYWORDS):
            return True
        if any(keyword in clean_desc for keyword in cls.NOISE_DESCRIPTION_KEYWORDS):
            return True
        if any(keyword in clean_sap_code for keyword in cls.NOISE_SAP_CODE_KEYWORDS) and any(
            keyword in clean_desc for keyword in cls.NOISE_DESCRIPTION_KEYWORDS
        ):
            return True
        return False

    @staticmethod
    def _generate_batch_id() -> str:
        """生成 BOM 导入批次号。"""
        return f"plan-bom-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:8]}"


__all__ = ["PlanBomExcelImportService"]
