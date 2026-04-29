from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from backend.app.db.base import Base
from backend.app.domains.plan_bom.constants import MATERIAL_CATEGORY_GLASS, STATUS_FAILED, STATUS_SUCCESS
from backend.app.domains.plan_bom.models import PlanBomHeader, PlanBomImportBatch, PlanBomMaterialLine, PlanBomRevision
from backend.app.domains.plan_bom.repositories.import_repository import PlanBomImportRepository
from backend.app.domains.plan_bom.services.excel_import_service import PlanBomExcelImportService


def _build_session() -> Session:
    """创建内存数据库会话。

    返回：
        已完成建表的 SQLAlchemy Session。
    """
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(bind=engine)
    session_factory = sessionmaker(bind=engine)
    return session_factory()


def _workbook_bytes(
    *,
    conflict: bool = False,
    order_no: str = "GCL-TEST-001",
    version_no: str = "A1",
    order_name: str = "测试订单",
    glass_description: str = "玻璃规格描述",
    include_gap_film: bool = False,
) -> bytes:
    """构造计划 BOM 测试 Excel。

    参数：
        conflict: 是否生成同一 `订单号 + 版本号 + SAP编码` 但内容不同的冲突行。

    返回：
        xlsx 二进制内容。
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    sheet.append(["文件号", "FILE-001", "版本号", version_no])
    sheet.append(["订单号", order_no, "订单名称", order_name])
    sheet.append([])
    sheet.append(["修订版本", "修订内容", "修订人", "生效日期"])
    sheet.append(["A1", "首版发布", "刘娟", "2026-04-01"])
    sheet.append([])
    sheet.append(["序号", "SAP编码", "物料名称", "描述", "标准用量", "单位", "生产损耗", "备注"])
    sheet.append([1, "1001", "光伏玻璃", glass_description, 2, "片", "0.01%", ""])
    sheet.append([2, "1002", "互联条", "焊带规格描述", 60, "根", "0.02%", ""])
    if include_gap_film:
        sheet.append([3, "1003", "间隙膜", "间隙膜规格描述", 4, "卷", "0.02%", ""])
    if conflict:
        sheet.append([4, "1001", "光伏玻璃", "另一条不同规格", 3, "片", "0.01%", ""])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _real_layout_workbook_bytes() -> bytes:
    """构造接近真实版式的计划 BOM Excel。

    说明：
        本测试样本用于覆盖两类真实问题：
        1. 材料区之后存在备注区、修订区和图纸区；
        2. 修订区位于材料区下方时，旧逻辑会直接跳过。
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "材料清单"
    sheet.append(["BOM", "", "", "", "", "", "文件号", "FILE-REAL-001"])
    sheet.append(["", "BOM", "", "", "", "", "版本号", "A1"])
    sheet.append(["", "1001", "", "", "", "", "订单号", "GCL-REAL-001"])
    sheet.append(["", "规格", "", "", "", "", "订单名称", "真实版式订单"])
    sheet.append(["序号", "SAP编码", "物料名称", "描述", "标准用量", "单位", "生产损耗", "备注"])
    sheet.append(["", "", "", "", "Qty.", "", "", ""])
    sheet.append([1, "SAP-GLASS", "光伏玻璃", "光伏玻璃\\彩虹\\有价值", 2, "片", "0.01%", ""])
    sheet.append([2, "SAP-JBOX", "接线盒", "接线盒\\谐通\\有价值", 1, "个", "0.01%", ""])
    sheet.append(['注：辅材搭配必须按照备注说明来搭配', '注：①材料名称后带"-A"表示选择其一', "", "", "", "", "", ""])
    sheet.append(["", "备注", "焊带汇流条更新为GCL", "", "", "", "", ""])
    sheet.append(["修订版本", "", "", "修订内容", "", "修订人", "生效日期", "备注"])
    sheet.append(["", "A1", "", "真实版式修订说明", "", "刘娟", "2026-04-01", ""])
    sheet.append(["编制/日期", "", "", "审核/日期", "会签/日期", "", "", "批准/日期"])
    sheet.append(["", "文控文件号", "文控版本", "图纸名称", "系统图号", "系统版本", "图纸顺序", ""])
    sheet.append(["", "GCL/XXJC/2-RD-0001", "A1", "GCL标准物料标签-接线盒 A1", "CD-001", "A", "10", ""])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _revision_fallback_workbook_bytes() -> bytes:
    """构造仅有修订表头和审批日期、没有显式修订明细的测试 Excel。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "材料清单"
    sheet.append(["文件号", "FILE-FALLBACK-001", "版本号", "A0"])
    sheet.append(["订单号", "GCL-FALLBACK-001", "订单名称", "审批日期兜底订单"])
    sheet.append(["序号", "SAP编码", "物料名称", "描述", "标准用量", "单位", "生产损耗", "备注"])
    sheet.append([1, "1001", "光伏玻璃", "光伏玻璃\\旗滨\\有价值", 2, "片", "0.01%", ""])
    sheet.append(["备注", "国内标准分布式，全0.24，370胶膜，需求功率630", "", "", "", "", "", ""])
    sheet.append(["修订版本", "", "", "修订内容", "", "修订人", "生效日期", "备注"])
    sheet.append(["编制/日期", "", "", "审核/日期", "会签/日期", "", "", "批准/日期"])
    sheet.append(["", "樊 娜/2026-03-20", "", "韩 小磊/2026-03-18", "", "", "", ""])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _usage_variant_workbook_bytes() -> bytes:
    """构造同一 SAP 编码仅用量不同的真实选配测试 Excel。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "材料清单"
    sheet.append(["文件号", "FILE-USAGE-001", "版本号", "A2"])
    sheet.append(["订单号", "GCL-USAGE-001", "订单名称", "用量变体订单"])
    sheet.append(["修订版本", "修订内容", "修订人", "生效日期"])
    sheet.append(["A2", "导入用量变体样本", "刘娟", "2026-04-18"])
    sheet.append(["序号", "SAP编码", "物料名称", "描述", "标准用量", "单位", "生产损耗", "备注"])
    sheet.append([1, "1000458780", "互联条 1", "互联条\\GCL\\φ0.24mm\\6040\\常规\\-\\有价值", 194.2, "千克", "0.10%", ""])
    sheet.append([2, "1000458780", "互联条", "互联条\\GCL\\φ0.24mm\\6040\\常规\\-\\有价值", 22, "千克", "0.10%", ""])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_plan_bom_excel_import_success() -> None:
    """验证 Excel 入库成功时会写入批次、BOM 头、材料行和修订区。"""
    session = _build_session()
    service = PlanBomExcelImportService(repository=PlanBomImportRepository(session))

    report = service.import_bytes(_workbook_bytes(), file_name="bom.xlsx", batch_id="batch-success")

    assert report.status == STATUS_SUCCESS
    assert report.header_count == 1
    assert report.material_line_count == 2
    assert report.revision_count == 1
    assert report.error_count == 0
    assert session.query(PlanBomImportBatch).count() == 1
    assert session.query(PlanBomHeader).count() == 1
    assert session.query(PlanBomMaterialLine).count() == 2
    assert session.query(PlanBomRevision).count() == 1

    glass_line = session.query(PlanBomMaterialLine).filter_by(sap_code="1001").one()
    assert glass_line.material_category == MATERIAL_CATEGORY_GLASS
    assert glass_line.source_tag == "manual_import_source"


def test_plan_bom_excel_import_detects_material_key_conflict() -> None:
    """验证失败批次会整批回滚业务数据，只保留批次记录。"""
    session = _build_session()
    service = PlanBomExcelImportService(repository=PlanBomImportRepository(session))

    report = service.import_bytes(_workbook_bytes(conflict=True), file_name="bom.xlsx", batch_id="batch-conflict")

    assert report.status == STATUS_FAILED
    assert report.error_count == 1
    assert report.rollback_applied is True
    assert report.persisted_business_data is False
    assert report.errors[0].stage == "MATERIAL_CONFLICT"
    assert "GCL-TEST-001|A1|1001|EXCEL" == report.errors[0].key
    assert session.query(PlanBomImportBatch).count() == 1
    assert session.query(PlanBomHeader).count() == 0
    assert session.query(PlanBomMaterialLine).count() == 0
    assert session.query(PlanBomRevision).count() == 0


def test_plan_bom_excel_import_real_layout_stops_material_section_and_parses_revision() -> None:
    """验证真实版式下材料区会在备注区前停止，并能解析材料区下方的修订区。"""
    session = _build_session()
    service = PlanBomExcelImportService(repository=PlanBomImportRepository(session))

    report = service.import_bytes(_real_layout_workbook_bytes(), file_name="real-layout.xlsx", batch_id="batch-real-layout")

    assert report.status == STATUS_SUCCESS
    assert report.rollback_applied is False
    assert report.persisted_business_data is True
    assert report.header_count == 1
    assert report.material_line_count == 2
    assert report.revision_count == 1

    header = session.query(PlanBomHeader).one()
    revision = session.query(PlanBomRevision).one()
    assert str(header.effective_date) == "2026-04-01"
    assert session.query(PlanBomRevision).count() == 1
    assert revision.revision_version == "A1"
    assert session.query(PlanBomMaterialLine).count() == 2
    assert session.query(PlanBomMaterialLine).filter_by(sap_code="备注").count() == 0


def test_plan_bom_excel_import_keeps_distinct_excel_instances_for_same_business_key() -> None:
    """验证 Excel 开发期内部实例键不会让同订单号同版本的不同实例互相覆盖。"""
    session = _build_session()
    service = PlanBomExcelImportService(repository=PlanBomImportRepository(session))

    first_report = service.import_bytes(
        _workbook_bytes(order_no="GCL-XXJC-JSPS-2026-00106", version_no="A0", order_name="江苏汉腾"),
        file_name="NT1078GDF(江苏汉腾-2026-00106)Billofmaterials-A.xls",
        batch_id="batch-00106-jsht",
    )
    second_report = service.import_bytes(
        _workbook_bytes(order_no="GCL-XXJC-JSPS-2026-00106", version_no="A0", order_name="石家庄科林"),
        file_name="NT1078GDF(石家庄科林-2026-00106)Billofmaterials-A.xls",
        batch_id="batch-00106-sjzkl",
    )

    assert first_report.status == STATUS_SUCCESS
    assert second_report.status == STATUS_SUCCESS
    headers = session.query(PlanBomHeader).order_by(PlanBomHeader.order_name.asc()).all()
    assert len(headers) == 2
    assert {header.order_name for header in headers} == {"江苏汉腾", "石家庄科林"}
    assert headers[0].order_identity_key != headers[1].order_identity_key
    assert session.query(PlanBomMaterialLine).count() == 4
    assert session.query(PlanBomRevision).count() == 2


def test_plan_bom_excel_import_keeps_multiple_file_instances_under_same_identity() -> None:
    """验证同一业务实例同版本下的不同文件实例可以并存，重复导入同一文件实例只覆盖自己。"""
    session = _build_session()
    service = PlanBomExcelImportService(repository=PlanBomImportRepository(session))

    first_report = service.import_bytes(
        _workbook_bytes(
            order_no="GCL-XXJC-JSPS-2026-00120",
            version_no="A0",
            order_name="肯尼亚 Nationwide",
            glass_description="光伏玻璃\\GCL\\有价值",
            include_gap_film=True,
        ),
        file_name="NT12R66GDF(肯尼亚Nationwide-2026-00120)Billofmaterials-A (2).xls",
        batch_id="batch-00120-a2",
    )
    second_report = service.import_bytes(
        _workbook_bytes(
            order_no="GCL-XXJC-JSPS-2026-00120",
            version_no="A0",
            order_name="肯尼亚 Nationwide",
            glass_description="光伏玻璃\\信义\\有价值",
            include_gap_film=False,
        ),
        file_name="NT12R66GDF(肯尼亚Nationwide-2026-00120)Billofmaterials-A (3).xls",
        batch_id="batch-00120-a3",
    )
    repeat_report = service.import_bytes(
        _workbook_bytes(
            order_no="GCL-XXJC-JSPS-2026-00120",
            version_no="A0",
            order_name="肯尼亚 Nationwide",
            glass_description="光伏玻璃\\GCL\\有价值",
            include_gap_film=True,
        ),
        file_name="NT12R66GDF(肯尼亚Nationwide-2026-00120)Billofmaterials-A (2).xls",
        batch_id="batch-00120-a2-repeat",
    )

    assert first_report.status == STATUS_SUCCESS
    assert second_report.status == STATUS_SUCCESS
    assert repeat_report.status == STATUS_SUCCESS

    headers = (
        session.query(PlanBomHeader)
        .filter(PlanBomHeader.order_no == "GCL-XXJC-JSPS-2026-00120")
        .order_by(PlanBomHeader.raw_file_name.asc())
        .all()
    )
    assert len(headers) == 2
    assert len({header.order_identity_key for header in headers}) == 1
    assert len({header.file_instance_key for header in headers}) == 2
    assert {header.raw_file_name for header in headers} == {
        "NT12R66GDF(肯尼亚Nationwide-2026-00120)Billofmaterials-A (2).xls",
        "NT12R66GDF(肯尼亚Nationwide-2026-00120)Billofmaterials-A (3).xls",
    }
    assert session.query(PlanBomMaterialLine).filter_by(order_no="GCL-XXJC-JSPS-2026-00120").count() == 5


def test_plan_bom_excel_import_falls_back_to_approval_date_when_revision_row_missing() -> None:
    """验证修订表头后只有审批区时，仍会用审批日期回填修订记录和生效日期。"""
    session = _build_session()
    service = PlanBomExcelImportService(repository=PlanBomImportRepository(session))

    report = service.import_bytes(
        _revision_fallback_workbook_bytes(),
        file_name="revision-fallback.xlsx",
        batch_id="batch-revision-fallback",
    )

    assert report.status == STATUS_SUCCESS
    assert report.revision_count == 1
    header = session.query(PlanBomHeader).one()
    revision = session.query(PlanBomRevision).one()
    assert str(header.effective_date) == "2026-03-20"
    assert revision.revision_version == "A0"
    assert revision.revision_content == "国内标准分布式，全0.24，370胶膜，需求功率630"
    assert revision.reviser == "樊 娜"


def test_plan_bom_excel_import_merges_usage_variants_for_same_sap_code() -> None:
    """验证同一 SAP 编码仅用量不同的材料行不会再导致整批失败。"""
    session = _build_session()
    service = PlanBomExcelImportService(repository=PlanBomImportRepository(session))

    report = service.import_bytes(
        _usage_variant_workbook_bytes(),
        file_name="usage-variant.xlsx",
        batch_id="batch-usage-variant",
    )

    assert report.status == STATUS_SUCCESS
    assert report.error_count == 0
    assert report.warning_count == 1
    assert report.warnings[0].stage == "MATERIAL_USAGE_VARIANT"
    material_line = session.query(PlanBomMaterialLine).one()
    assert str(material_line.standard_usage) == "194.200000"
