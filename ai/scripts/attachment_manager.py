#!/usr/bin/env python3
"""AI 附件管理：复制附件、生成附件清单和摘要。

设计原则：附件先变成 TASK 资产，再交给 Hermes 技术经理和 Codex 引用；原始附件默认只读，不让 Codex
直接覆盖。
"""
from __future__ import annotations

import csv
import json
import shutil
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

PROJECT_ROOT = Path(__file__).resolve().parents[2]
INBOX_DIR = PROJECT_ROOT / "ai" / "inbox"
INBOX_ATTACHMENTS = INBOX_DIR / "attachments"
INBOX_MANIFEST = INBOX_DIR / "attachments_manifest.md"

TEXT_SUFFIXES = {".md", ".txt", ".csv", ".json", ".yaml", ".yml", ".sql", ".py", ".ts", ".vue"}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
DOCX_SUFFIXES = {".docx"}
PDF_SUFFIXES = {".pdf"}
ZIP_SUFFIXES = {".zip"}


def copy_inbox_attachments(task_dir: Path) -> list[Path]:
    """复制 ai/inbox/attachments 到 TASK 目录。

    返回值：复制后的附件文件列表。
    """
    copied: list[Path] = []
    if INBOX_ATTACHMENTS.exists():
        target = task_dir / "attachments"
        if target.exists():
            shutil.rmtree(target)
        shutil.copytree(INBOX_ATTACHMENTS, target)
        copied = [p for p in target.rglob("*") if p.is_file()]
    else:
        (task_dir / "attachments").mkdir(parents=True, exist_ok=True)

    if INBOX_MANIFEST.exists():
        shutil.copyfile(INBOX_MANIFEST, task_dir / "attachments_manifest.md")
    else:
        (task_dir / "attachments_manifest.md").write_text("# 附件清单\n\n当前未提供附件清单。\n", encoding="utf-8")

    return copied


def _safe_read_text(path: Path, max_chars: int = 6000) -> str:
    """安全读取文本附件，避免把大文件直接塞进 prompt。"""
    try:
        text = path.read_text(encoding="utf-8", errors="ignore")
    except Exception as exc:  # noqa: BLE001
        return f"读取失败：{exc}"
    return text[:max_chars] + ("\n...（已截断）" if len(text) > max_chars else "")


def _summarize_csv(path: Path, max_rows: int = 10) -> str:
    """生成 CSV 的表头和样例行摘要。"""
    try:
        with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as f:
            reader = csv.reader(f)
            rows = []
            for idx, row in enumerate(reader):
                rows.append(row)
                if idx >= max_rows:
                    break
        if not rows:
            return "CSV 为空。"
        return "表头/样例行：\n" + "\n".join([" | ".join(r) for r in rows])
    except Exception as exc:  # noqa: BLE001
        return f"CSV 摘要失败：{exc}"


def _summarize_excel(path: Path) -> str:
    """生成 Excel 工作表、表头和前几行样例摘要。"""
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # noqa: BLE001
        return f"未安装 openpyxl，无法自动解析 Excel。请先人工转为 CSV 或安装 openpyxl。错误：{exc}"

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        lines = [f"工作表数量：{len(wb.sheetnames)}", f"工作表：{', '.join(wb.sheetnames)}"]
        for ws in wb.worksheets[:5]:
            lines.append(f"\n### Sheet: {ws.title}")
            lines.append(f"最大行列：rows={ws.max_row}, cols={ws.max_column}")
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 8), values_only=True):
                values = ["" if v is None else str(v) for v in row[:20]]
                lines.append(" | ".join(values))
        return "\n".join(lines)
    except Exception as exc:  # noqa: BLE001
        return f"Excel 摘要失败：{exc}"


def _summarize_docx(path: Path) -> str:
    """用标准库从 docx 中抽取部分文本。"""
    try:
        with zipfile.ZipFile(path) as zf:
            xml = zf.read("word/document.xml")
        root = ET.fromstring(xml)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        texts = [node.text or "" for node in root.findall(".//w:t", ns)]
        text = "".join(texts)
        return text[:6000] + ("\n...（已截断）" if len(text) > 6000 else "")
    except Exception as exc:  # noqa: BLE001
        return f"docx 摘要失败：{exc}"


def _summarize_zip(path: Path) -> str:
    """只列出 zip 文件树，不自动解压覆盖项目。"""
    try:
        with zipfile.ZipFile(path) as zf:
            names = zf.namelist()[:200]
        return "压缩包文件列表（前 200 个）：\n" + "\n".join(names)
    except Exception as exc:  # noqa: BLE001
        return f"zip 摘要失败：{exc}"


def summarize_attachment(path: Path, task_dir: Path) -> str:
    """按类型生成单个附件摘要。"""
    rel = path.relative_to(task_dir)
    suffix = path.suffix.lower()
    size = path.stat().st_size
    header = [f"## {rel}", "", f"- 类型后缀：{suffix or '无'}", f"- 文件大小：{size} bytes", ""]

    if suffix == ".csv":
        body = _summarize_csv(path)
    elif suffix in TEXT_SUFFIXES:
        body = _safe_read_text(path)
    elif suffix in EXCEL_SUFFIXES:
        body = _summarize_excel(path)
    elif suffix in DOCX_SUFFIXES:
        body = _summarize_docx(path)
    elif suffix in PDF_SUFFIXES:
        body = "PDF 附件已登记。V1 不自动 OCR / 不强行解析 PDF；如需精读，请先转为 md/txt 或提供摘要。"
    elif suffix in IMAGE_SUFFIXES:
        body = "图片附件已登记。V1 仅记录文件路径；请在 attachments_manifest.md 中补充页面/截图描述。"
    elif suffix in ZIP_SUFFIXES:
        body = _summarize_zip(path)
    else:
        body = "未知类型附件已登记。V1 不自动解析，请在 attachments_manifest.md 中补充用途说明。"

    return "\n".join(header) + body + "\n"


def build_attachments_summary(task_dir: Path) -> Path:
    """生成 attachments_summary.md 和 attachments_inventory.json。"""
    attachments_dir = task_dir / "attachments"
    attachments_dir.mkdir(parents=True, exist_ok=True)
    files = [p for p in attachments_dir.rglob("*") if p.is_file()]

    inventory = []
    sections = [
        "# 附件摘要",
        "",
        "说明：本文件由 ai/scripts/attachment_manager.py 自动生成。原始附件默认只读，Codex 应优先读取本摘要和 attachments_manifest.md。",
        "",
    ]
    if not files:
        sections.append("当前任务没有附件。\n")

    for path in files:
        inventory.append({
            "path": str(path.relative_to(task_dir)),
            "suffix": path.suffix.lower(),
            "size_bytes": path.stat().st_size,
        })
        sections.append(summarize_attachment(path, task_dir))

    summary_file = task_dir / "attachments_summary.md"
    summary_file.write_text("\n".join(sections), encoding="utf-8")
    (task_dir / "attachments_inventory.json").write_text(json.dumps(inventory, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary_file
