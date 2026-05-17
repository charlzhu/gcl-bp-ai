from pathlib import Path
from openpyxl import load_workbook
import json
import re
from collections import Counter

root = Path('/Users/zhuchangchao/Work/PythonProject/project/gcl-bp-ai')
task = root / 'ai/tasks/running/TASK-business-feedback-excel-qa-fix'
task.mkdir(parents=True, exist_ok=True)
excel = root / '经营计划智能体测试统计.xlsx'

QUESTION_KEYWORDS = ('问题', '提问', '题目', '用户问题', '业务员提问', '测试问题', 'prompt', 'Prompt')
ANSWER_KEYWORDS = ('回答', '答案', '系统回答', 'AI回答', '实际答案', '结果', '回复')
FEEDBACK_KEYWORDS = ('反馈', '备注', '错误', '原因', '业务员备注', '期望', '正确', '问题点', '说明', '审核')
STATUS_KEYWORDS = ('状态', '是否', '已解答', '修复')

def sval(value):
    """把单元格内容转换为稳定文本。"""
    if value is None:
        return ''
    text = str(value).replace('\r', '\n').strip()
    return re.sub(r'\n{3,}', '\n\n', text)

def short(text, limit=260):
    """生成 Markdown 中使用的单行摘要。"""
    text = sval(text).replace('\n', ' / ')
    return text[:limit] + ('…' if len(text) > limit else '')

def classify_domain(text):
    """粗分业务域，后续复现阶段会再用服务实际结果校正。"""
    if any(keyword in text for keyword in ('订单', 'BOM', 'bom', '功率', '焊带', '玻璃', '汇流条', '线盒', '线长', '电池效率', '供应商', '间隙贴膜')):
        return 'plan_bom'
    if any(keyword in text for keyword in ('物流', '发运', '运费', '运量', '车次', '承运', '区域', 'MW', '单瓦', '线路', '运输', '经营计划', '派车', '任务')):
        return 'logistics'
    return 'unknown'

def classify_issue(text):
    """按业务员反馈文本粗分缺陷类型。"""
    tags = []
    if any(keyword in text for keyword in ('图表', '表格', '可视化', '曲线', '柱状', '折线', '下载', '导出')):
        tags.append('presentation')
    if any(keyword in text for keyword in ('值', '不准', '错误', '不对', '应该', '不是', '差', '少', '多', '没有', '漏')):
        tags.append('value_or_retrieval')
    if any(keyword in text for keyword in ('结构', '臃肿', '清晰', '啰嗦', '简洁')):
        tags.append('answer_structure')
    if any(keyword in text for keyword in ('澄清', '反问', '追问', '补充')):
        tags.append('clarification')
    return tags or ['needs_manual_triage']

wb = load_workbook(excel, read_only=True, data_only=False)
items = []
lines = []
lines.append('# 业务反馈 Excel 解析')
lines.append('')
lines.append(f'- 文件：`{excel}`')
lines.append(f'- 存在：{excel.exists()}')
lines.append(f'- 大小：{excel.stat().st_size if excel.exists() else 0} bytes')
lines.append(f'- Sheet：{", ".join(wb.sheetnames)}')
lines.append('')

for ws in wb.worksheets:
    rows = list(ws.iter_rows(values_only=True))
    non_empty = [[sval(cell) for cell in row] for row in rows if any(sval(cell) for cell in row)]
    lines.append(f'## Sheet: {ws.title}')
    lines.append(f'- 原始范围：{ws.max_row} 行 x {ws.max_column} 列')
    lines.append(f'- 非空行：{len(non_empty)}')
    if not non_empty:
        lines.append('')
        continue

    header_idx = 0
    best_score = -1
    for idx, row in enumerate(non_empty[:10]):
        joined = ' '.join(row)
        score = sum(1 for keyword in QUESTION_KEYWORDS + ANSWER_KEYWORDS + FEEDBACK_KEYWORDS + STATUS_KEYWORDS if keyword.lower() in joined.lower()) + sum(1 for cell in row if cell)
        if score > best_score:
            header_idx = idx
            best_score = score

    headers = [cell if cell else f'列{idx + 1}' for idx, cell in enumerate(non_empty[header_idx])]
    lines.append(f'- 识别表头行：非空行 #{header_idx + 1}')
    lines.append('- 表头：' + ' | '.join(f'{idx + 1}:{header}' for idx, header in enumerate(headers)))

    question_cols = []
    answer_cols = []
    feedback_cols = []
    status_cols = []
    for idx, header in enumerate(headers):
        header_lower = header.lower()
        if any(keyword.lower() in header_lower for keyword in QUESTION_KEYWORDS):
            question_cols.append(idx)
        if any(keyword.lower() in header_lower for keyword in ANSWER_KEYWORDS):
            answer_cols.append(idx)
        if any(keyword.lower() in header_lower for keyword in FEEDBACK_KEYWORDS):
            feedback_cols.append(idx)
        if any(keyword.lower() in header_lower for keyword in STATUS_KEYWORDS):
            status_cols.append(idx)
    if not question_cols:
        question_cols = [next((idx for idx, _ in enumerate(headers) if idx > 0), 0)]

    lines.append(f'- 猜测问题列：{[headers[idx] for idx in question_cols]}')
    lines.append(f'- 猜测回答列：{[headers[idx] for idx in answer_cols]}')
    lines.append(f'- 猜测反馈/备注列：{[headers[idx] for idx in feedback_cols]}')
    lines.append('')

    for row_no, row in enumerate(non_empty[header_idx + 1:], start=header_idx + 2):
        row = row + [''] * (len(headers) - len(row))
        question = '\n'.join(row[idx] for idx in question_cols if idx < len(row) and row[idx]).strip()
        answer = '\n'.join(row[idx] for idx in answer_cols if idx < len(row) and row[idx]).strip()
        feedback = '\n'.join(row[idx] for idx in feedback_cols if idx < len(row) and row[idx]).strip()
        status_text = '\n'.join(row[idx] for idx in status_cols if idx < len(row) and row[idx]).strip()
        all_text = '\n'.join(row)
        if not (question or feedback or answer):
            continue
        if len(question) < 2 and not feedback:
            continue

        all_cells = {headers[idx] if idx < len(headers) else f'列{idx + 1}': row[idx] for idx in range(min(len(headers), len(row))) if row[idx]}
        item = {
            'id': f'{ws.title}-R{row_no}',
            'sheet': ws.title,
            'excel_row': row_no,
            'question': question,
            'answer': answer,
            'feedback': feedback,
            'status_text': status_text,
            'all_cells': all_cells,
        }
        item['domain'] = classify_domain(question + '\n' + feedback + '\n' + answer)
        item['issue_tags'] = classify_issue(question + '\n' + feedback + '\n' + answer)
        item['has_answered_text'] = '已解答' in all_text
        items.append(item)

        excluded = {headers[idx] for idx in question_cols + answer_cols + feedback_cols + status_cols if idx < len(headers)}
        extras = {key: value for key, value in all_cells.items() if key not in excluded}
        lines.append(f"### {item['id']}")
        lines.append(f"- 领域：{item['domain']}；标签：{', '.join(item['issue_tags'])}；含“已解答”字样：{item['has_answered_text']}")
        lines.append(f"- 问题：{short(question, 500)}")
        if feedback:
            lines.append(f"- 反馈/备注：{short(feedback, 500)}")
        if answer:
            lines.append(f"- 现有回答摘录：{short(answer, 500)}")
        if extras:
            lines.append(f"- 其他列：{json.dumps({key: short(value, 160) for key, value in extras.items()}, ensure_ascii=False)}")
        lines.append('')

lines.append('## 解析统计')
lines.append(f'- 提取问题/反馈行数：{len(items)}')
lines.append(f'- 按领域：{dict(Counter(item["domain"] for item in items))}')
lines.append(f'- 含“已解答”字样行数：{sum(1 for item in items if item["has_answered_text"])}')

(task / 'excel-analysis.md').write_text('\n'.join(lines), encoding='utf-8')
(task / 'excel-items.json').write_text(json.dumps(items, ensure_ascii=False, indent=2, default=str), encoding='utf-8')
(task / 'task.md').write_text(
    "# TASK-business-feedback-excel-qa-fix\n\n"
    "## 目标\n"
    "基于 `经营计划智能体测试统计.xlsx` 中业务员反馈的不准问答，逐条核对问题值、答案结构、图表/表格使用和表达清晰度，定位通用根因并按 TDD 修复。\n\n"
    "## 边界\n"
    "- 允许：后端问答 planner/service/repository、前端 business-chat 展示、业务验收测试、任务验收材料。\n"
    "- 禁止：自动 commit/push/deploy；修改 .env/密钥；数据库迁移；硬编码单题答案；覆盖原始 Excel。\n\n"
    "## 验收标准\n"
    "1. Excel 行不能因“已解答”字样跳过，必须复现或说明数据/口径不可自动确认原因。\n"
    "2. 通用修复覆盖同类问法，新增 RED/GREEN 回归测试。\n"
    "3. focused/related/full/backend compile/frontend build/static scan/reviewer 通过或说明阻断。\n"
    "4. 生成 diff.patch、test.log、review_bundle.md、final-acceptance.md。\n",
    encoding='utf-8',
)
print(json.dumps({'task_dir': str(task), 'items': len(items), 'domains': dict(Counter(item['domain'] for item in items))}, ensure_ascii=False))
